from django.shortcuts import render,redirect
from django.http import HttpResponseRedirect
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from .models import *
from .forms import *
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.contrib import messages
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from Users.models import AuditTrail  # Import AuditTrail model
from django.utils.timezone import now  # For timestamping
from django.shortcuts import render,redirect, get_object_or_404
import pandas as pd
from django.views import View
from django.contrib import messages
from threading import Thread
from queue import Queue
from django.db import transaction, IntegrityError, DatabaseError
from django.core.exceptions import ValidationError
from django.db import connection
from django.apps import apps
from django.forms import modelform_factory
from django.db.models import Q
import csv
from django.http import HttpResponse,JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import IntegrityError, DatabaseError as DBError
from django.views import View
import pandas as pd
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.http import HttpResponse
from django.http import JsonResponse
import requests
from django.contrib.auth.decorators import login_required
from Users.models import AuditTrail  # Import AuditTrail model
from django.utils.timezone import now  # For timestamping
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from Users.models import AuditTrail  # Import AuditTrail model
from django.utils.timezone import now  # For timestamping
from django.shortcuts import render, get_object_or_404, redirect
from django.forms import inlineformset_factory
from django.db import transaction
import threading
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from Users.models import AuditTrail  # Import AuditTrail model
from django.utils.timezone import now  # For timestamping
from django.db import transaction
from django.db.models import Q
from django.db.models import Count
from django.http import JsonResponse
from django.template.loader import render_to_string
from django.db.models import Max
from django.db.models import Min
from datetime import datetime
from Users.models import AuditTrail  # Import AuditTrail model
from django.utils.timezone import now  # For timestamping
from django.utils import timezone
from django.utils.module_loading import import_string  # Used for dynamic function calling
from django.core.paginator import Paginator
from openpyxl.utils import get_column_letter
import csv
import openpyxl
from django.http import JsonResponse, HttpResponse
from IFRS9.signals import *
from django.db.models import Min, Max
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.db import IntegrityError
from Users.models import AuditTrail  # Import AuditTrail model
from django.utils.timezone import now  # For timestamping
from ast import Import
from django.shortcuts import render,redirect, get_object_or_404
from django.db.models import Max
from django.contrib import messages
import csv
from django.http import HttpResponse
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Q
from django.contrib.auth.decorators import login_required
import pandas as pd
import numpy as np
from datetime import datetime
import os
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Max, Sum, Count, Q
from django.shortcuts import render,redirect
from django.http import HttpResponseRedirect
from django.urls import reverse
import matplotlib.pyplot as plt
import io
import  base64
from Users.models import AuditTrail  # Import AuditTrail model
from django.utils.timezone import now  # For timestamping
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.contrib.auth.decorators import login_required
from django.urls import reverse_lazy
from django.core.paginator import Paginator
from django.contrib import messages
from django.db import transaction
from django.views.decorators.http import require_http_methods
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.urls import reverse_lazy
from django.db import IntegrityError
from django.views.generic import UpdateView
from django.contrib.auth.mixins import LoginRequiredMixin
from IFRS9.signals import *







def dashboard_view(request):
    # Example data for financial graphs
    mis_date = '2024-08-31'  # Input date in 'YYYY-MM-DD' format
    #perform_interpolation(mis_date) 
   

    return render(request, 'dashboard.html')



def app_list_view(request):
    context = {
        'title': 'Available Applications',
        # You can pass any additional context if needed
    }
    return render(request, 'app_list.html', context)

@login_required
def ifrs9_home_view(request):
    context = {
        'title': ' Home',
        # You can pass any additional context if needed
    }
    return render(request, 'ifrs9_home.html', context)

@login_required
def credit_risk_models_view(request):
    context = {
        'title': 'Credit Risk Models',
    }
    return render(request, 'models/credit_risk_models.html', context)



#####################
def cashflow_projection_view(request):
    if request.method == 'POST':
        fic_mis_date = request.POST.get('fic_mis_date')
        
        # Trigger the projection
        project_cash_flows(fic_mis_date)
        
        # Redirect to a success page
        return HttpResponseRedirect(reverse('projection_success'))
    
    return render(request, 'cashflow_projection.html')

@login_required
def cashflow_projections(request):
    # This view will render the page that shows two options: Documentation and Interest Method
    context = {
        'title': 'Cashflow Projections',
        # No need to pass the URLs from the view, as they're now hardcoded in the HTML
    }
    return render(request, 'cashflow_projections/index.html', context)

@login_required
def cashflow_projections_documentation(request):
    # You can pass any context data if needed
    context = {
        'title': 'Cash Flow Generation Issues and Solutions',
    }
    return render(request, 'cashflow_projections/cash_flow_generation_issues.html', context)



# List View

class InterestMethodListView(LoginRequiredMixin,ListView):
    model = Fsi_Interest_Method
    template_name = 'cashflow_projections/interest_method_list.html'
    context_object_name = 'methods'

# Create View
class InterestMethodCreateView(LoginRequiredMixin,CreateView):
    model = Fsi_Interest_Method
    form_class = InterestMethodForm
    template_name = 'cashflow_projections/interest_method_form.html'
    success_url = reverse_lazy('interest_method_list')

    def form_valid(self, form):
        # Set the created_by field to the currently logged-in user
        instance = form.save(commit=False)
        instance.created_by = self.request.user
        instance.save()   
        # Log the creation in the AuditTrail
        AuditTrail.objects.create(
                user=self.request.user,
                model_name='Fsi_Interest_Method',
                action='create',
                object_id=instance.pk,
                change_description=f"Created Interest Method: {instance.v_interest_method}",
                timestamp=now(),
            )  
        messages.success(self.request, "Interest method added successfully!")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "There was an error adding the interest method. Please try again.")
        return super().form_invalid(form)

# Update View
class InterestMethodUpdateView(LoginRequiredMixin,UpdateView):
    model = Fsi_Interest_Method
    form_class = InterestMethodForm
    template_name = 'cashflow_projections/interest_method_form.html'
    success_url = reverse_lazy('interest_method_list')

    def form_valid(self, form):
        # Set the created_by field to the currently logged-in user
        instance = form.save(commit=False)
        previous_method = self.get_object().v_interest_method 
        instance.created_by = self.request.user
        instance.save() 
        # Log the update in the AuditTrail
        AuditTrail.objects.create(
                user=self.request.user,
                model_name='Fsi_Interest_Method',
                action='update',
                object_id=instance.pk,
                change_description=(
                    f"Updated Interest Method: From {previous_method} to {instance.v_interest_method}"
                ),
                timestamp=now(),
            )    
        messages.success(self.request, "Interest method updated successfully!")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "There was an error updating the interest method. Please try again.")
        return super().form_invalid(form)

# Delete View
class InterestMethodDeleteView(LoginRequiredMixin, DeleteView):
    model = Fsi_Interest_Method
    template_name = 'cashflow_projections/interest_method_confirm_delete.html'
    success_url = reverse_lazy('interest_method_list')

    def delete(self, request, *args, **kwargs):
        instance = self.get_object()
        try:
            # Log the deletion in the AuditTrail
            AuditTrail.objects.create(
                user=self.request.user,
                model_name='Fsi_Interest_Method',
                action='delete',
                object_id=instance.pk,
                change_description=f"Deleted Interest Method: {instance.v_interest_method}",
                timestamp=now(),
            )
            messages.success(self.request, "Interest method deleted successfully!")
        except Exception as e:
            messages.error(self.request, f"An unexpected error occurred while logging the deletion: {e}")
        return super().delete(request, *args, **kwargs)
    

@login_required
def data_management(request):
    return render(request, 'load_data/data_management.html')

class FileUploadView(LoginRequiredMixin,View):
    template_name = 'load_data/file_upload_step1.html'

    def get(self, request):
        form = UploadFileForm()

        # Fetch staging tables from the database
        stg_tables = TableMetadata.objects.filter(table_type='STG').values_list('table_name', flat=True)

        return render(request, self.template_name, {
            'form': form,
            'stg_tables': stg_tables  # Pass the staging tables to the template
        })

    def post(self, request):
        form = UploadFileForm(request.POST, request.FILES)
        selected_table = request.POST.get('table_name')  # Get selected table from the form

        if form.is_valid() and selected_table:
            file = form.cleaned_data['file']

            try:
                # Automatically detect file type and process accordingly
                if file.name.endswith('.csv'):
                    df = pd.read_csv(file)
                elif file.name.endswith(('.xls', '.xlsx')):
                    df = pd.read_excel(file)
                else:
                    messages.error(request, "Unsupported file format. Please upload a CSV or Excel file.")
                    return render(request, self.template_name, {'form': form, 'stg_tables': TableMetadata.objects.filter(table_type='STG')})

                # Convert Timestamps to strings for JSON serialization
                df = df.applymap(lambda x: x.isoformat() if isinstance(x, pd.Timestamp) else x)


                # Store the data, column names, and selected table in session for later steps
                request.session['file_data'] = df.to_dict()  # Save the full data in session
                request.session['columns'] = list(df.columns)
                request.session['selected_table'] = selected_table  # Save the selected table

                # Prepare preview for rendering (first 10 rows)
                preview_data = {
                    'headers': list(df.columns),
                    'rows': df.head(10).values.tolist()  # Show the first 10 rows for preview
                }

                return render(request, self.template_name, {
                    'form': form,
                    'preview_data': preview_data,
                    'show_next_button': True,
                    'table_name': selected_table,
                    'file_name': file.name,
                    'stg_tables': TableMetadata.objects.filter(table_type='STG')
                })
            except Exception as e:
                messages.error(request, f"Error processing file: {e}")
        return render(request, self.template_name, {
            'form': form,
            'stg_tables': TableMetadata.objects.filter(table_type='STG')
        })
    
class ColumnSelectionView(LoginRequiredMixin,View):
    template_name = 'load_data/file_upload_step2.html'

    def get(self, request):
        columns = request.session.get('columns', [])
        if not columns:
            messages.error(request, "No columns found. Please upload a file first.")
            return redirect('file_upload')
        form = ColumnSelectionForm(columns=columns, initial={'selected_columns': columns})
        return render(request, self.template_name, {'form': form, 'columns': columns})

    def post(self, request):
        selected_columns = request.POST.get('selected_columns_hidden').split(',')
        if selected_columns:
            request.session['selected_columns'] = selected_columns
            return redirect('map_columns')
        else:
            messages.error(request, "You must select at least one column.")
        return render(request, self.template_name, {'form': form})

########################
class ColumnMappingView(LoginRequiredMixin,View):
    template_name = 'load_data/file_upload_step3.html'

    def get(self, request):
        selected_columns = request.session.get('selected_columns', [])
        selected_table = request.session.get('selected_table')  # Get the selected table from the session

        # Get the model class dynamically based on the selected table
        try:
            model_class = apps.get_model('IFRS9', selected_table)  # Replace 'IFRS9' with your actual app name
        except LookupError:
            messages.error(request, "Error: The selected table does not exist.")
            return render(request, self.template_name)

        model_fields = [f.name for f in model_class._meta.fields]

        # Create initial mappings based on matching names (case-insensitive)
        initial_mappings = {}
        unmapped_columns = []

        for column in selected_columns:
            match = next((field for field in model_fields if field.lower() == column.lower()), None)
            if match:
                initial_mappings[column] = match
            else:
                initial_mappings[column] = 'unmapped'  # Set to 'unmapped' if no match is found
                unmapped_columns.append(column)  # Track unmapped columns

        # Initialize the form with the mappings
        form = ColumnMappingForm(
            initial={'column_mappings': initial_mappings}, 
            selected_columns=selected_columns, 
            model_fields=model_fields
        )

  

        # Check if there are unmapped columns
        if unmapped_columns:
            messages.warning(request, "The following columns were not automatically mapped: " + ", ".join(unmapped_columns))
        
        return render(request, self.template_name, {'form': form, 'unmapped_columns': unmapped_columns})

    def post(self, request):
        selected_columns = request.session.get('selected_columns', [])
        selected_table = request.session.get('selected_table')  # Get the selected table from the session

        # Get the model class dynamically based on the selected table
        try:
            model_class = apps.get_model('IFRS9', selected_table)  # Replace 'IFRS9' with your actual app name
        except LookupError:
            messages.error(request, "Error: The selected table does not exist.")
            return render(request, self.template_name)

        model_fields = [f.name for f in model_class._meta.fields]

        # Initialize the form with POST data
        form = ColumnMappingForm(request.POST, selected_columns=selected_columns, model_fields=model_fields)

        if form.is_valid():
            # Safely get the 'column_mappings' from cleaned_data
            mappings = form.cleaned_data.get('column_mappings', {})

           

            # Validate that all columns have been mapped (i.e., not mapped to 'unmapped')
            unmapped_columns = [col for col, mapped_to in mappings.items() if mapped_to == 'unmapped']
            if unmapped_columns:
                messages.error(request, "The following columns are not mapped: " + ", ".join(unmapped_columns))
                return render(request, self.template_name, {'form': form, 'unmapped_columns': unmapped_columns})

            # Ensure that there are mappings before proceeding
            if not mappings or all(value == 'unmapped' for value in mappings.values()):
                messages.error(request, "Error: No valid column mappings provided. Please map all columns before proceeding.")
                return render(request, self.template_name, {'form': form, 'unmapped_columns': unmapped_columns})

            # Save the mappings to the session
            request.session['column_mappings'] = mappings

            return redirect('submit_to_database')

        # If the form is not valid, render the form again with errors
        messages.error(request, "Error: Invalid form submission. Please check your mappings and try again.")
        return render(request, self.template_name, {'form': form})


#####################




class SubmitToDatabaseView(LoginRequiredMixin, View):
    template_name = 'load_data/file_upload_step4.html'

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        try:
            # Retrieve data from the session
            df_data = request.session.get('file_data')
            selected_columns = request.session.get('selected_columns')
            mappings = request.session.get('column_mappings')
            selected_table = request.session.get('selected_table')

            if not df_data or not selected_columns or not mappings:
                return JsonResponse({'status': 'error', 'message': 'Missing data in the session.'}, status=400)

            # Convert session data to DataFrame and apply mappings
            df = pd.DataFrame(df_data)
            df = df[selected_columns].rename(columns=mappings)

            # Clean the data
            df = self.clean_data(df)

            # Convert date columns to YYYY-MM-DD format
            for column in df.columns:
                if 'date' in column.lower():
                    df[column] = pd.to_datetime(df[column], errors='coerce').dt.strftime('%Y-%m-%d').astype(str)
            
            # Retrieve the model class dynamically
            model_class = apps.get_model('IFRS9', selected_table)

            # Define chunk size for bulk insert
            chunk_size = 5000
            total_chunks = len(df) // chunk_size + (1 if len(df) % chunk_size > 0 else 0)
            request.session['progress'] = 0

            # Counter to track the total number of records successfully inserted
            success_count = 0

            for i, chunk_start in enumerate(range(0, len(df), chunk_size), start=1):
                # Prepare the chunk for insertion
                chunk_df = df.iloc[chunk_start:chunk_start + chunk_size]
                records_to_insert = [model_class(**row.to_dict()) for _, row in chunk_df.iterrows()]

                try:
                    # Bulk create the records with conflict handling
                    model_class.objects.bulk_create(records_to_insert)
                    success_count += len(records_to_insert)  # Increment success count by inserted records
                except IntegrityError as e:
                    return JsonResponse({'status': 'error', 'message': f'Integrity error: {str(e)}'}, status=400)
                except ValidationError as e:
                    error_messages = "; ".join(f"{field}: {', '.join(errors)}" for field, errors in e.message_dict.items())
                    return JsonResponse({'status': 'error', 'message': f'Validation error: {error_messages}'}, status=400)
                except DBError as e:
                    return JsonResponse({'status': 'error', 'message': f'Database error: {str(e)}'}, status=400)

                # Update progress in session after each chunk
                request.session['progress'] = int((i / total_chunks) * 100)
                request.session.modified = True

            # Mark completion and return success message with count of records uploaded
            request.session['progress'] = 100
            return JsonResponse({'status': 'success', 'message': f'{success_count} records successfully uploaded.'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f"Unexpected error: {str(e)}"}, status=500)
        
    def clean_data(self, df):
        """
        Cleans the DataFrame by applying various data cleaning steps.
        """
        # Remove leading and trailing spaces from all string columns
        for column in df.select_dtypes(include=['object']).columns:
            df[column] = df[column].str.strip()

        # Trim whitespace and normalize spaces between words
        for column in df.select_dtypes(include=['object']).columns:
            df[column] = df[column].str.strip().str.replace(r'\s+', ' ', regex=True)

        # Convert all text to uppercase
        for column in df.select_dtypes(include=['object']).columns:
            df[column] = df[column].str.upper()
        # Additional cleaning steps can be added here, such as:
        df.dropna(how='all', inplace=True)  # Drop rows where all values are NaN
        # Drop duplicates
        df.drop_duplicates(inplace=True)

        return df
class CheckProgressView(LoginRequiredMixin, View):
    def get(self, request):
        progress = request.session.get('progress', 0)
        return JsonResponse({'progress': progress})
    

     
    ####################################################################
@login_required
def data_entry_view(request):
    table_form = TableSelectForm(request.POST or None)
    data_form = None

    if request.method == 'POST':
        if table_form.is_valid():
            selected_table = table_form.cleaned_data['table_name'].table_name  # Get the selected table's name

            try:
                # Get the model class dynamically
                model_class = apps.get_model('IFRS9', selected_table)  # Replace 'IFRS9' with your actual app name
            except LookupError:
                messages.error(request, "Error: The selected table does not exist.")
                return render(request, 'load_data/data_entry.html', {'table_form': table_form, 'data_form': data_form})

            # Dynamically create a form for the selected model
            DynamicForm = modelform_factory(model_class, fields='__all__')
            data_form = DynamicForm(request.POST or None)

            if data_form.is_valid():
                try:
                    data_form.save()
                    messages.success(request, "Data successfully saved!")
                    return redirect('data_entry')
                except IntegrityError as e:
                    messages.error(request, f"Database Error: {e}")
                except ValidationError as e:
                    messages.error(request, f"Validation Error: {e.message_dict}")
                except Exception as e:
                    messages.error(request, f"Unexpected Error: {e}")
    
    return render(request, 'load_data/data_entry.html', {
        'table_form': table_form,
        'data_form': data_form
    })


@login_required
def get_fic_mis_dates(request):
    table_name = request.GET.get('table_name')
    try:
        model_class = apps.get_model('IFRS9', table_name)
        dates = model_class.objects.values_list('fic_mis_date', flat=True).distinct().order_by('-fic_mis_date')
        date_choices = [(date, date.strftime('%Y-%m-%d')) for date in dates]
        return JsonResponse({'dates': date_choices})
    except LookupError:
        return JsonResponse({'error': 'Table not found'}, status=404)
    
########################################################

class TableSelectForm(LoginRequiredMixin, forms.Form):
    table_name = forms.ChoiceField(choices=[], label="--select table--")
    fic_mis_date = forms.ChoiceField(choices=[('', '--select date--')], required=False, label="Select Date")

    def __init__(self, *args, **kwargs):
        # Get the initial table_name and fic_mis_date if they exist
        initial_table_name = kwargs['data'].get('table_name') if 'data' in kwargs else None
        initial_fic_mis_date = kwargs['data'].get('fic_mis_date') if 'data' in kwargs else None
        
        super().__init__(*args, **kwargs)
        
        # Populate table_name choices with a prompt
        app_models = apps.get_app_config('IFRS9').get_models()
        self.fields['table_name'].choices = [('', '--select table--')] + [
            (model._meta.model_name, model._meta.verbose_name) for model in app_models
        ]

        # If a table is selected, populate fic_mis_date choices
        if initial_table_name:
            model_class = apps.get_model('IFRS9', initial_table_name)
            distinct_dates = model_class.objects.values_list('fic_mis_date', flat=True).distinct().order_by('-fic_mis_date')
            self.fields['fic_mis_date'].choices = [('', '--select date--')] + [
                (date, date.strftime('%Y-%m-%d')) for date in distinct_dates
            ]
            # Set the initial value for fic_mis_date
            if initial_fic_mis_date:
                self.fields['fic_mis_date'].initial = initial_fic_mis_date

@login_required
def view_data(request):
    table_form = TableSelectForm(request.GET or None)
    data = None
    columns = []
    column_unique_values = {}
    table_name = request.GET.get('table_name')
    fic_mis_date = request.GET.get('fic_mis_date')

    if table_form.is_valid():
        if table_name and fic_mis_date:
            try:
                # Get the model class dynamically based on the selected table
                model_class = apps.get_model('IFRS9', table_name)
                print(f"Selected Table: {table_name}, Selected Date: {fic_mis_date}")

                # Apply filtering by `fic_mis_date` using `__date` for compatibility
                data = model_class.objects.filter(fic_mis_date__date=fic_mis_date)
                
                # Debug output
                print(f"Query: {data.query}")
                print(f"Data Count: {data.count()}")  # Check if data is returned

                # Get column names and unique values for each column
                columns = [field.name for field in model_class._meta.fields]
                for column in columns:
                    column_unique_values[column] = model_class.objects.values_list(column, flat=True).distinct()
                
                if not data:
                    messages.warning(request, "No data found for the selected date.")
                
            except LookupError:
                messages.error(request, "Error: The selected table does not exist.")
        else:
            messages.warning(request, "Please select both a table and a date to view data.")
            
            # Repopulate `fic_mis_date` choices based on selected table
            if table_name:
                model_class = apps.get_model('IFRS9', table_name)
                distinct_dates = model_class.objects.values_list('fic_mis_date', flat=True).distinct().order_by('-fic_mis_date')
                table_form.fields['fic_mis_date'].choices = [('', 'Select Date')] + [(date, date) for date in distinct_dates]

    return render(request, 'load_data/view_data.html', {
        'table_form': table_form,
        'data': data,
        'columns': columns,
        'column_unique_values': column_unique_values,
        'table_name': table_name,
        'fic_mis_date': fic_mis_date,
    })


@login_required
def filter_table(request):
    # Get table_name from the GET parameters
    table_name = request.GET.get('table_name')
    table_form = TableSelectForm(initial={'table_name': table_name})
    data = None
    columns = []
    column_unique_values = {}
    
    # Retrieve other parameters
    fic_mis_date = request.GET.get('fic_mis_date')
    filter_column = request.GET.get('filter_column')
    filter_values = request.GET.get('filter_values')
    sort_order = request.GET.get('sort_order')

    if table_name and fic_mis_date:  # Ensure both table_name and fic_mis_date are required for data retrieval
        try:
            model_class = apps.get_model('IFRS9', table_name)
            
            # Start by filtering directly on fic_mis_date, without __date lookup
            data = model_class.objects.filter(fic_mis_date=fic_mis_date)

            # Additional column-specific filtering if provided
            if filter_column and filter_values:
                filter_values_list = filter_values.split(',')
                filter_values_list = [value for value in filter_values_list if value not in ["on", "(Select All)"]]
                
                filters = Q()
                if "None" in filter_values_list:
                    filter_values_list.remove("None")
                    filters |= Q(**{f"{filter_column}__isnull": True})
                if filter_values_list:
                    filters |= Q(**{f"{filter_column}__in": filter_values_list})
                
                data = data.filter(filters)

            # Apply sorting if specified
            if filter_column and sort_order:
                data = data.order_by(filter_column if sort_order == 'asc' else f'-{filter_column}')

            # Retrieve columns and unique values for dropdown filters
            columns = [field.name for field in model_class._meta.fields]
            for column in columns:
                column_unique_values[column] = data.values_list(column, flat=True).distinct()

            # Debug output
            print(f"SQL Query: {data.query}")
            print(f"Data Count: {data.count()}")

        except LookupError:
            messages.error(request, "Error: The selected table does not exist.")
            print("Error: The selected table does not exist.")
    else:
        messages.warning(request, "Please select both a table and a date to view data.")

    return render(request, 'load_data/view_data.html', {
        'table_form': table_form,
        'data': data,
        'columns': columns,
        'column_unique_values': column_unique_values,
        'table_name': table_name,
        'fic_mis_date': fic_mis_date,
    })


@login_required
def download_data(request, table_name):
    try:
        # Get the model class dynamically using the table name
        model_class = apps.get_model('IFRS9', table_name)
        data = model_class.objects.all()

        # Handle filtering via GET parameters
        filter_column = request.GET.get('filter_column')
        filter_values = request.GET.get('filter_values')

        if filter_column and filter_values:
            filter_values_list = filter_values.split(',')

            # Filter out any unwanted values like "on" or "Select All"
            filter_values_list = [value for value in filter_values_list if value not in ["on", "(Select All)"]]

            # Prepare a Q object to combine multiple conditions
            filters = Q()

            # If "None" is selected, add an isnull filter
            if "None" in filter_values_list:
                filter_values_list.remove("None")
                filters |= Q(**{f"{filter_column}__isnull": True})

            # If there are other values selected, add the in filter
            if filter_values_list:
                filters |= Q(**{f"{filter_column}__in": filter_values_list})

            # Apply the combined filter to the data
            data = data.filter(filters)

        # Handle sorting via GET parameters
        sort_order = request.GET.get('sort_order')
        if filter_column and sort_order:
            if sort_order == 'asc':
                data = data.order_by(filter_column)
            elif sort_order == 'desc':
                data = data.order_by(f'-{filter_column}')

        # Create the HTTP response object with the appropriate CSV header.
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{table_name}.csv"'

        writer = csv.writer(response)
        # Write the headers
        writer.writerow([field.name for field in model_class._meta.fields])

        # Write the data rows
        for item in data:
            writer.writerow([getattr(item, field.name) for field in model_class._meta.fields])

        return response

    except LookupError:
        messages.error(request, "Error: The selected table does not exist.")
        return redirect('view_data')


@login_required
def edit_row(request, table_name, row_id):
    try:
        # Get the model class dynamically
        model_class = apps.get_model('IFRS9', table_name)
        row = get_object_or_404(model_class, id=row_id)

        if request.method == 'POST':
            # Update the row with new data from the AJAX request
            for field, value in request.POST.items():
                if field != 'csrfmiddlewaretoken':
                    setattr(row, field, value)
            row.save()

            return JsonResponse({'success': True})  # Respond with success

        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=400)

    except LookupError:
        return JsonResponse({'success': False, 'error': 'Table not found'}, status=404)

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def delete_row(request, table_name, row_id):
    try:
        # Get the model class dynamically
        model_class = apps.get_model('IFRS9', table_name)
        row = get_object_or_404(model_class, id=row_id)

        if request.method == 'POST':
            row.delete()
            return JsonResponse({'success': True})  # Respond with success

        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=400)

    except LookupError:
        return JsonResponse({'success': False, 'error': 'Table not found'}, status=404)

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)



@login_required
def ifrs9_configuration(request):
    # Render the IFRS9 Configuration template
    return render(request, 'ifrs9_conf/ifrs9_configuration.html')

@login_required
def ecl_methodology_options(request):
    # This view shows the two options: Documentation and Choose Methodology
    return render(request, 'ifrs9_conf/ecl_methodology_options.html')

@login_required
def ecl_methodology_documentation(request):
    # View to show the ECL methodology documentation
    return render(request, 'ifrs9_conf/ecl_methodology_documentation.html') 

 # You would create a separate documentation page
@login_required
def ecl_methodology_list(request):
    methods = ECLMethod.objects.all()
    return render(request, 'ifrs9_conf/ecl_methodology_list.html', {'methods': methods})

@login_required
def add_ecl_method(request):
    if request.method == 'POST':
        form = ECLMethodForm(request.POST)
        if form.is_valid():
            try:
                # Save the form without committing to the database yet
                ecl_method = form.save(commit=False)
                # Set the created_by field to the currently logged-in user
                ecl_method.created_by = request.user
                # Save the object to the database
                ecl_method.save()
                # Log the creation in the AuditTrail
                AuditTrail.objects.create(
                    user=request.user,
                    model_name='ECLMethod',
                    action='create',
                    object_id=ecl_method.pk,
                    change_description=f"Created ECL Methodology: {ecl_method.method_name}",
                    timestamp=now(),
                )
                messages.success(request, "New ECL Methodology added successfully!")
                return redirect('ecl_methodology_list')
            except ValidationError as e:
                form.add_error(None, e.message)  # This adds the validation error to the form's non-field errors
        else:
            messages.error(request, "There was an error adding the ECL Methodology.")
    else:
        form = ECLMethodForm()

    return render(request, 'ifrs9_conf/add_ecl_method.html', {'form': form})


@login_required
def edit_ecl_method(request, method_id):
    method = get_object_or_404(ECLMethod, pk=method_id)
    if request.method == 'POST':
        form = ECLMethodForm(request.POST, instance=method)
        if form.is_valid():
            try:
                # Save the form without committing to the database yet
                previous_values = {
                    "method_name": method.method_name,
                    "uses_discounting": method.uses_discounting,
                }
                ecl_method = form.save(commit=False)
                # Set the created_by field to the currently logged-in user
                ecl_method.created_by = request.user
                # Save the object to the database
                ecl_method.save()

                # Log the update in the AuditTrail
                AuditTrail.objects.create(
                    user=request.user,
                    model_name='ECLMethod',
                    action='update',
                    object_id=ecl_method.pk,
                    change_description=(
                        f"Updated ECL Methodology: From {previous_values} to "
                        f"{form.cleaned_data}"
                    ),
                    timestamp=now(),
                )

                messages.success(request, "ECL methodology updated successfully!")
                return redirect('ecl_methodology_list')
            except Exception as e:
                messages.error(request, f"An unexpected error occurred: {e}")
        else:
            messages.error(request, "There was an error updating the ECL methodology.")
    else:
        form = ECLMethodForm(instance=method)
    
    return render(request, 'ifrs9_conf/edit_ecl_method.html', {'form': form})


@login_required
def delete_ecl_method(request, method_id):
    method = get_object_or_404(ECLMethod, pk=method_id)
    
    if request.method == 'POST':
        try:
            # Log the deletion in the AuditTrail
            AuditTrail.objects.create(
                user=request.user,
                model_name='ECLMethod',
                action='delete',
                object_id=method.pk,
                change_description=f"Deleted ECL Methodology: {method.method_name}",
                timestamp=now(),
            )

            method.delete()
            messages.success(request, "ECL Methodology deleted successfully!")
            return redirect('ecl_methodology_list')
        except Exception as e:
            messages.error(request, f"An unexpected error occurred: {e}")

    return render(request, 'ifrs9_conf/delete_ecl_method.html', {'method': method})

@login_required
def choose_ecl_methodology(request):
    # View to configure the ECL methodology
    return render(request, 'ifrs9_conf/choose_ecl_methodology.html')  





@login_required
def lgd_configuration(request):
    return render(request, 'lgd_conf/lgd_configuration.html')



# List all LGD Term Structures
@login_required
def lgd_term_structure_list(request):
    term_structures = Ldn_LGD_Term_Structure.objects.all()
    return render(request, 'lgd_conf/lgd_term_structure_list.html', {'term_structures': term_structures})

# Create a new LGD Term Structure
@login_required
def lgd_term_structure_create(request):
    if request.method == 'POST':
        form = LGDTermStructureForm(request.POST)
        if form.is_valid():
            # Save the form without committing to the database yet
            term_structure = form.save(commit=False)
                # Set the created_by field to the currently logged-in user
            term_structure.created_by = request.user
                # Save the object to the database
            term_structure.save()
            # Log the creation in the AuditTrail
            AuditTrail.objects.create(
                    user=request.user,
                    model_name='Ldn_LGD_Term_Structure',
                    action='create',
                    object_id=term_structure.pk,
                    change_description=(f"Created LGD Term Structure: Name - {term_structure.v_lgd_term_structure_name}, "
                                        f"Description - {term_structure.v_lgd_term_structure_desc}, LGD% - {term_structure.n_lgd_percent}"),
                    timestamp=now(),
                )

            messages.success(request, "LGD Term Structure added successfully!")
            return redirect('lgd_term_structure_list')
        else:
            messages.error(request, "Error adding LGD Term Structure.")
    else:
        form = LGDTermStructureForm()
    
    return render(request, 'lgd_conf/lgd_term_structure_form.html', {'form': form})

# Edit LGD Term Structure
@login_required
def lgd_term_structure_edit(request, term_id):
    term_structure = get_object_or_404(Ldn_LGD_Term_Structure, pk=term_id)
    if request.method == 'POST':
        form = LGDTermStructureForm(request.POST, instance=term_structure)
        if form.is_valid():
            term_structure = form.save(commit=False)
                # Set the created_by field to the currently logged-in user
            term_structure.created_by = request.user
                # Save the object to the database
            term_structure.save()
            # Log the update in the AuditTrail
            AuditTrail.objects.create(
                    user=request.user,
                    model_name='Ldn_LGD_Term_Structure',
                    action='update',
                    object_id=term_structure.pk,
                    change_description=(f"Updated LGD Term Structure: Name - {term_structure.v_lgd_term_structure_name}, "
                                        f"Description - {term_structure.v_lgd_term_structure_desc}, LGD% - {term_structure.n_lgd_percent}"),
                    timestamp=now(),
                )
          
            messages.success(request, "LGD Term Structure updated successfully!")
            return redirect('lgd_term_structure_list')
        else:
            messages.error(request, "Error updating LGD Term Structure.")
    else:
        form = LGDTermStructureForm(instance=term_structure)
    
    return render(request, 'lgd_conf/lgd_term_structure_form.html', {'form': form})

# Delete LGD Term Structure
@login_required
def lgd_term_structure_delete(request, term_id):
    term_structure = get_object_or_404(Ldn_LGD_Term_Structure, pk=term_id)
    if request.method == 'POST':
        # Log the deletion in the AuditTrail
        AuditTrail.objects.create(
                user=request.user,
                model_name='Ldn_LGD_Term_Structure',
                action='delete',
                object_id=term_structure.pk,
                change_description=(f"Deleted LGD Term Structure: Name - {term_structure.v_lgd_term_structure_name}, "
                                    f"Description - {term_structure.v_lgd_term_structure_desc}, LGD% - {term_structure.n_lgd_percent}"),
                timestamp=now(),
            )
        term_structure.delete()
        messages.success(request, "LGD Term Structure deleted successfully!")
        return redirect('lgd_term_structure_list')
    return render(request, 'lgd_conf/lgd_term_structure_confirm_delete.html', {'term_structure': term_structure})



@login_required
def view_lgd_calculation(request):
    # Try to retrieve the single CollateralLGD instance
    lgd_instance = CollateralLGD.objects.first()

    # If no instance exists, create one with default values
    if not lgd_instance:
        try:
            lgd_instance = CollateralLGD.objects.create(can_calculate_lgd=False)
        except ValidationError as e:
            return render(request, 'lgd_conf/view_lgd_calculation.html', {'error': str(e)})

    # Pass the instance to the template
    return render(request, 'lgd_conf/view_lgd_calculation.html', {'lgd_instance': lgd_instance})

@login_required
def edit_lgd_calculation(request):
    """Edit the LGD Calculation settings"""
    # Retrieve the first instance of CollateralLGD, or show a 404 if none exists
    lgd_instance = CollateralLGD.objects.first()

    if not lgd_instance:
        messages.error(request, "No LGD Calculation data available to edit.")
        return redirect('view_lgd_calculation')  # Redirect to a safer page

    if request.method == 'POST':
        form = CollateralLGDForm(request.POST, instance=lgd_instance)
        if form.is_valid():
            form.save()
            messages.success(request, "LGD Calculation settings updated successfully!")
            return redirect('view_lgd_calculation')
        else:
            messages.error(request, "Error updating LGD Calculation settings.")
    else:
        form = CollateralLGDForm(instance=lgd_instance)

    return render(request, 'lgd_conf/edit_lgd_calculation.html', {'form': form})


@login_required
def operations_view(request):
    return render(request, 'operations/operations.html')



# List all processes
@login_required
def process_list(request):
    processes = Process.objects.all()
    return render(request, 'operations/process_list.html', {'processes': processes})

# View the details of a specific process, including its associated functions and execution order
@login_required
def process_detail(request, process_id):
    process = get_object_or_404(Process, id=process_id)
    run_processes = RunProcess.objects.filter(process=process).order_by('order')  # Fetch functions in order
    return render(request, 'operations/process_detail.html', {'process': process, 'run_processes': run_processes})



##############################################################################################3
# Display and search for processes
@login_required
def execute_process_view(request):
    query = request.GET.get('search', '')
    processes = Process.objects.filter(Q(process_name__icontains=query))

    return render(request, 'operations/execute_process.html', {
        'processes': processes,
        'query': query,
    })

# Handle execution
# Function to generate the process run ID and count

running_threads = {}
cancel_flags = {}


def generate_process_run_id(process, execution_date):
    """
    Generate a process_run_id in the format 'process_id_execution_date_run_number'.
    """
    # Format the execution date as YYYYMMDD
    execution_date_str = execution_date.strftime('%Y%m%d')
    
    # Base run ID: process_id + execution_date
    base_run_id = f"{process.process_name}_{execution_date_str}"
    
    # Check the database for existing entries with the same base_run_id
    existing_runs = FunctionExecutionStatus.objects.filter(process_run_id__startswith=base_run_id).order_by('-run_count')
    
    # Determine the next run count
    if existing_runs.exists():
        last_run_count = existing_runs[0].run_count
        next_run_count = last_run_count + 1
    else:
        next_run_count = 1
    
    # Generate the full process_run_id with the next run count
    process_run_id = f"{base_run_id}_{next_run_count}"
    
    return process_run_id, next_run_count


# Background function for running the process
# Background function for running the process
def execute_functions_in_background(function_status_entries, process_run_id, mis_date):
    for status_entry in function_status_entries:

        if cancel_flags.get(process_run_id):  # Check if cancellation was requested
            status_entry.status = 'Cancelled'
            status_entry.execution_end_date = timezone.now()
            status_entry.duration = status_entry.execution_end_date - status_entry.execution_start_date
            status_entry.save()
            print(f"Process {process_run_id} was cancelled.")
            break  # Stop execution if cancelled

        function_name = status_entry.function.function_name
        print(f"Preparing to execute function: {function_name}")

        # Set the function status to "Ongoing" and record the start date
        status_entry.status = 'Ongoing'
        status_entry.execution_start_date = timezone.now()  # Start time for the function
        status_entry.save()
        print(f"Function {function_name} marked as Ongoing.")

        # Execute the function
        try:
            if function_name in globals():
                print(f"Executing function: {function_name} with date {mis_date}")
                result = globals()[function_name](mis_date)  # Execute the function and capture the return value
                
                # Update status and end date based on the return value
                status_entry.execution_end_date = timezone.now()  # End time for the function
                if result == 1 or result == '1':
                    status_entry.status = 'Success'
                    print(f"Function {function_name} executed successfully.")
                elif result == 0 or result == '0':
                    status_entry.status = 'Failed'
                    print(f"Function {function_name} execution failed.")
                    status_entry.save()
                    break  # Stop execution if the function fails
                else:
                    status_entry.status = 'Failed'
                    print(f"Unexpected return value {result} from function {function_name}.")
                    status_entry.save()
                    break  # Stop execution for any unexpected result
            else:
                status_entry.status = 'Failed'
                print(f"Function {function_name} not found in the global scope.")
                status_entry.execution_end_date = timezone.now()
                status_entry.save()
                break  # Stop execution if the function is not found

        except Exception as e:
            status_entry.status = 'Failed'
            status_entry.execution_end_date = timezone.now()
            print(f"Error executing {function_name}: {e}")
            status_entry.save()
            break  # Stop execution if any function throws an exception

        # Calculate duration
        if status_entry.execution_start_date and status_entry.execution_end_date:
            status_entry.duration = status_entry.execution_end_date - status_entry.execution_start_date

        # Save the final status and duration
        status_entry.save()
        print(f"Updated FunctionExecutionStatus for {function_name} to {status_entry.status}")


running_threads = {}
cancel_flags = {}

@login_required
def run_process_execution(request):
    if request.method == 'POST':
        process_id = request.POST.get('process_id')
        mis_date = request.POST.get('execution_date')

        if not process_id or not mis_date:
            messages.error(request, "Please select a process and execution date.")
            return redirect('execute_process_view')

        # Convert execution date string to date object
        execution_date = datetime.strptime(mis_date, '%Y-%m-%d')

        # Fetch selected process
        process = get_object_or_404(Process, id=process_id)

        # Fetch related functions in execution order
        run_processes = RunProcess.objects.filter(process=process).order_by('order')

        if not run_processes.exists():
            messages.error(request, "No functions are associated with this process.")
            return redirect('execute_process_view')

        # Generate process_run_id
        process_run_id, run_count = generate_process_run_id(process, execution_date)

        # Save function execution statuses
        function_status_entries = []
        for run_process in run_processes:
            status_entry = FunctionExecutionStatus.objects.create(
                process=process,
                function=run_process.function,
                reporting_date=mis_date,
                status='Pending',
                process_run_id=process_run_id,
                run_count=run_count,
                execution_order=run_process.order,
                created_by=request.user
            )
            function_status_entries.append(status_entry)

        # Redirect to monitoring page immediately
        response = redirect('monitor_specific_process', process_run_id=process_run_id)

        # Start function execution in a background thread
        execution_thread = threading.Thread(
            target=execute_functions_in_background,
            args=(function_status_entries, process_run_id, mis_date)
        )
        execution_thread.start()

        messages.success(request, f"Process '{process.process_name}' started successfully!")
        return response

    return redirect('execute_process_view')


@login_required
def get_process_functions(request, process_id):
    process = get_object_or_404(Process, id=process_id)
    functions_html = render_to_string('operations/_functions_list.html', {'run_processes': process.run_processes.all()})
    return JsonResponse({'html': functions_html})



# Monitor running process page
@login_required
def monitor_running_process_view(request):
    # Fetch distinct reporting dates and order by date descending
    available_dates = FunctionExecutionStatus.objects.order_by('-reporting_date').values_list('reporting_date', flat=True).distinct()

    # Get the selected date from the request
    selected_date = request.GET.get('selected_date', '')

    # Filter processes based on the selected date and ensure uniqueness by using annotation
    processes = []
    if selected_date:
        processes = (
            FunctionExecutionStatus.objects.filter(reporting_date=selected_date)
            .values('process__process_name', 'process_run_id','created_by__email')
            .annotate(
                latest_run=Max('process_run_id'),  # Latest run ID
                start_time=Min('execution_start_date'),  # Earliest start time
                end_time=Max('execution_end_date')  # Latest end time
            )
        )

        # Calculate overall status and duration for each process
        for process in processes:
            process_run_id = process['process_run_id']
            function_statuses = FunctionExecutionStatus.objects.filter(process_run_id=process_run_id)

            # Determine the overall status based on the function statuses
            if function_statuses.filter(status='Failed').exists():
                process['overall_status'] = 'Failed'
            elif function_statuses.filter(status='Cancelled').exists():
                process['overall_status'] = 'Cancelled'
            elif function_statuses.filter(status='Ongoing').exists():
                process['overall_status'] = 'Ongoing'
            else:
                process['overall_status'] = 'Success'
            

            # Calculate duration if start and end times are available
            start_time = process['start_time']
            end_time = process['end_time']
            if start_time and end_time:
                duration = end_time - start_time
                process['duration'] = duration.total_seconds() / 60  # Convert to minutes
            else:
                process['duration'] = None

    context = {
        'selected_date': selected_date,
        'processes': processes,
        'available_dates': available_dates,

    }
    return render(request, 'operations/monitor_running_process.html', context)


@login_required
def monitor_specific_process(request, process_run_id):
    # Fetch the specific process run by its ID
    process_statuses = FunctionExecutionStatus.objects.filter(process_run_id=process_run_id)

    context = {
        'process_statuses': process_statuses,
        'process_run_id': process_run_id,
    }
    return render(request, 'operations/monitor_specific_process.html', context)

@login_required
def get_updated_status_table(request):
    process_run_id = request.GET.get('process_run_id')
    
    # Get all statuses related to the process_run_id
    function_statuses = FunctionExecutionStatus.objects.filter(process_run_id=process_run_id).order_by('execution_order')
    
    # Add total_seconds to each status entry
    for status in function_statuses:
        if isinstance(status.duration, timedelta):
            # Calculate total seconds and assign it as an additional attribute for template access
            status.total_seconds = status.duration.total_seconds()
        else:
            status.total_seconds = None

    # Return the partial template with the updated table
    return render(request, 'operations/status_table.html', {'function_statuses': function_statuses})

@login_required
def get_process_function_status(request, process_run_id):
    # Get all statuses related to the process_run_id
    run_processes = FunctionExecutionStatus.objects.filter(process_run_id=process_run_id).order_by('execution_order')

    # Compute duration details for each process
    for process in run_processes:
        if isinstance(process.duration, timedelta):
            total_seconds = process.duration.total_seconds()
            process.total_seconds = total_seconds
            process.minutes = int(total_seconds // 60)
            process.remaining_seconds = int(total_seconds % 60)
        else:
            process.total_seconds = None
            process.minutes = None
            process.remaining_seconds = None

    # Render the updated table to HTML
    functions_html = render_to_string('operations/_function_status_list.html', {'run_processes': run_processes})
    return JsonResponse({'html': functions_html})



@login_required
def running_processes_view(request):
    # Fetch all ongoing (running) processes from the FunctionExecutionStatus model
    running_processes = FunctionExecutionStatus.objects.filter(status='Ongoing')

    context = {
        'running_processes': running_processes
    }

    return render(request, 'operations/running_processes.html', context)

# Updated function to handle cancellation request
@login_required
def cancel_running_process(request, process_run_id):
    # Check if the process is running
    try:
        # Get all functions related to the given process_run_id
        functions = FunctionExecutionStatus.objects.filter(
            process_run_id=process_run_id,
            status__in=['Pending', 'Ongoing']
        )
        
        if functions.exists():
            # Update the status of all "Pending" and "Ongoing" functions to "Cancelled"
            functions.update(status='Cancelled')
            messages.success(request, f"Process '{process_run_id}' and all pending functions have been cancelled.")
        else:
            messages.info(request, f"No running process found with the given ID '{process_run_id}'.")
    
    except FunctionExecutionStatus.DoesNotExist:
        messages.error(request, f"An error occurred while cancelling the process '{process_run_id}'.")

    return redirect('running_processes')  # Redirect to the running processes list




@login_required
def probability_configuration(request):
    return render(request, 'probability_conf/probability_configuration.html')

@login_required
def pd_modelling(request):
    # Logic for PD Modelling
    return render(request, 'probability_conf/pd_modelling.html')


# List all segments with pagination
@login_required
def segment_list(request):
    segments = FSI_Product_Segment.objects.all()
    rows_per_page = int(request.GET.get('rows', 5))  # Default to 8 rows per page
    paginator = Paginator(segments, rows_per_page)

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Calculate the number of empty rows to add if there are fewer than `rows_per_page` on the page
    empty_rows = rows_per_page - len(page_obj)

    return render(request, 'probability_conf/segment_list.html', {
        'page_obj': page_obj,
        'rows_per_page': rows_per_page,
        'empty_rows': range(empty_rows)  # Pass a range object to loop through empty rows in the template
    })


# Create new segment
@login_required
def segment_create(request):
    if request.method == "POST":
        form = FSIProductSegmentForm(request.POST)
        if form.is_valid():
            # Save the form without committing to the database yet
            segment = form.save(commit=False)
            # Set the created_by field to the currently logged-in user
            segment.created_by = request.user
            # Save the object to the database
            segment.save()

            # Log the creation in the AuditTrail
            AuditTrail.objects.create(
                    user=request.user,
                    model_name='FSI_Product_Segment',
                    action='create',
                    object_id=segment.pk,
                    change_description=(
                        f"Created Segment: Segment - {segment.v_prod_segment}, "
                        f"Type - {segment.v_prod_type}, Description - {segment.v_prod_desc}"
                    ),
                    timestamp=now(),
                )
            
            messages.success(request, "Segment added successfully!")
            return redirect('segment_list')
        else:
            error_message = form.errors.as_json()
            messages.error(request, f"There was an error in adding the segment: {error_message}")
    else:
        form = FSIProductSegmentForm()
    return render(request, 'probability_conf/segment_form.html', {'form': form})

@login_required
def get_product_types(request):
    segment = request.GET.get('segment')
    if segment:
        types = Ldn_Bank_Product_Info.objects.filter(v_prod_segment=segment).values_list('v_prod_type', flat=True).distinct()
        type_choices = [{'value': t, 'display': t} for t in types]
    else:
        type_choices = []
    return JsonResponse(type_choices, safe=False)

@login_required
def get_product_description(request):
    product_type = request.GET.get('product_type')
    if product_type:
        description = Ldn_Bank_Product_Info.objects.filter(v_prod_type=product_type).values_list('v_prod_desc', flat=True).first()
        return JsonResponse({'description': description})
    return JsonResponse({'description': ''})


# Update existing segment
@login_required
def segment_edit(request, segment_id):
    segment = get_object_or_404(FSI_Product_Segment, pk=segment_id)
    
    if request.method == "POST":
        form = FSIProductSegmentForm(request.POST, instance=segment)
        if form.is_valid():
            try:
                # Save the form without committing to the database yet
                segment = form.save(commit=False)
                # Set the created_by field to the currently logged-in user
                segment.created_by = request.user
                # Save the object to the database
                segment.save()
                #form.save()
                # Log the update in the AuditTrail
                AuditTrail.objects.create(
                    user=request.user,
                    model_name='FSI_Product_Segment',
                    action='update',
                    object_id=segment.pk,
                    change_description=(
                        f"Updated Segment: Segment - {segment.v_prod_segment}, "
                        f"Type - {segment.v_prod_type}, Description - {segment.v_prod_desc}"
                    ),
                    timestamp=now(),
                )
                messages.success(request, "Segment updated successfully!")
                return redirect('segment_list')
            except IntegrityError as e:
                # Handle duplicate entry or unique constraint errors
                messages.error(request, f"Integrity error: {e}")
            except Exception as e:
                # Handle any other unexpected errors
                messages.error(request, f"An unexpected error occurred: {e}")
        else:
            # General message for validation errors; specific errors will show in the form
            messages.error(request, "There were errors in the form. Please correct them below.")
    else:
        form = FSIProductSegmentForm(instance=segment)
    
    return render(request, 'probability_conf/segment_form.html', {'form': form})

# Delete segment
@login_required
def segment_delete(request, segment_id):
    segment = get_object_or_404(FSI_Product_Segment, pk=segment_id)
    if request.method == "POST":
        # Log the deletion in the AuditTrail
        AuditTrail.objects.create(
            user=request.user,
            model_name='FSI_Product_Segment',
            action='delete',
            object_id=segment.pk,
            change_description=(
                f"Deleted Segment: Segment - {segment.v_prod_segment}, "
                f"Type - {segment.v_prod_type}, Description - {segment.v_prod_desc}"
            ),
            timestamp=now(),
        )

        segment.delete()
        messages.success(request, "Segment deleted successfully!")
        return redirect('segment_list')
    return render(request, 'probability_conf/segment_confirm_delete.html', {'segment': segment})

@login_required
def pd_term_structure_list(request):
    term_structures_list = Ldn_PD_Term_Structure.objects.all()
    
    # Set up pagination
    paginator = Paginator(term_structures_list, 5)  # Show 5 items per page
    page_number = request.GET.get('page')
    term_structures = paginator.get_page(page_number)  # Get the page of items

    return render(request, 'probability_conf/pd_term_structure_list.html', {'term_structures': term_structures})

# Create new PD Term Structure
@login_required
def pd_term_structure_create(request):
    if request.method == "POST":
        form = PDTermStructureForm(request.POST)
        if form.is_valid():
            try:
                # Save the form without committing to the database yet
                term_structure = form.save(commit=False)
                # Set the created_by field to the currently logged-in user
                term_structure.created_by = request.user
                # Save the object to the database
                term_structure.save()
                # Log the creation in the AuditTrail
                AuditTrail.objects.create(
                    user=request.user,
                    model_name='Ldn_PD_Term_Structure',
                    action='create',
                    object_id=term_structure.pk,
                    change_description=(
                        f"Created PD Term Structure: Name - {term_structure.v_pd_term_structure_name}, "
                        f"Date - {term_structure.fic_mis_date}"
                    ),
                    timestamp=now(),
                )

                
                messages.success(request, "PD Term Structure added successfully!")
                return redirect('pd_term_structure_list')
            except IntegrityError as e:
                # Handle duplicate entry error with the specific exception message
                messages.error(request, f"Error adding PD Term Structure: {e}")
            except Exception as e:
                # Handle any other exceptions
                messages.error(request, f"An unexpected error occurred: {e}")
        else:
            # Display form validation errors
            messages.error(request, "There were errors in the form. Please correct them below.")
    else:
        form = PDTermStructureForm()

    return render(request, 'probability_conf/pd_term_structure_form.html', {'form': form})

@login_required
def pd_term_structure_edit(request, term_id):
    term_structure = get_object_or_404(Ldn_PD_Term_Structure, pk=term_id)
    
    if request.method == "POST":
        form = PDTermStructureForm(request.POST, instance=term_structure)
        if form.is_valid():
            try:
                # Save the form without committing to the database yet
                term_structure = form.save(commit=False)
                # Set the created_by field to the currently logged-in user
                term_structure.created_by = request.user
                # Save the object to the database
                term_structure.save()
                # Log the update in the AuditTrail
                AuditTrail.objects.create(
                    user=request.user,
                    model_name='Ldn_PD_Term_Structure',
                    action='update',
                    object_id=term_structure.pk,
                    change_description=(
                        f"Updated PD Term Structure: Name - {term_structure.v_pd_term_structure_name}, "
                        f"Date - {term_structure.fic_mis_date}"
                    ),
                    timestamp=now(),
                )

                messages.success(request, "PD Term Structure updated successfully!")
                return redirect('pd_term_structure_list')
            except IntegrityError as e:
                # Handle duplicate entry error with the specific exception message
                messages.error(request, f"Error updating PD Term Structure: {e}")
            except Exception as e:
                # Catch any other exceptions and show the error message
                messages.error(request, f"An unexpected error occurred: {e}")
        else:
            messages.error(request, "There were validation errors. Please correct them below.")
    else:
        form = PDTermStructureForm(instance=term_structure)
    
    return render(request, 'probability_conf/pd_term_structure_form.html', {'form': form})
# Delete PD Term Structure
@login_required
def pd_term_structure_delete(request, term_id):
    term_structure = get_object_or_404(Ldn_PD_Term_Structure, pk=term_id)
    if request.method == "POST":
        # Log the deletion in the AuditTrail
        AuditTrail.objects.create(
            user=request.user,
            model_name='Ldn_PD_Term_Structure',
            action='delete',
            object_id=term_structure.pk,
            change_description=(
                f"Deleted PD Term Structure: Name - {term_structure.v_pd_term_structure_name}, "
                f"Date - {term_structure.fic_mis_date}"
            ),
            timestamp=now(),
        )
        term_structure.delete()
        messages.success(request, "PD Term Structure deleted successfully!")
        return redirect('pd_term_structure_list')
    return render(request, 'probability_conf/pd_term_structure_confirm_delete.html', {'term_structure': term_structure})


####################################33
# List all Delinquent Based PD Terms
@login_required
def delinquent_pd_list(request):
    # Filter for 'Delinquent' type PD Term Structures
    pd_term_details_list = Ldn_PD_Term_Structure_Dtl.objects.filter(
        v_pd_term_structure_id__v_pd_term_structure_type='D'
    ).select_related('v_pd_term_structure_id')
    
    # Set up pagination with 5 items per page
    paginator = Paginator(pd_term_details_list, 5)
    page_number = request.GET.get('page')
    pd_term_details = paginator.get_page(page_number)

    return render(request, 'probability_conf/delinquent_pd_list.html', {'pd_term_details': pd_term_details})

# Create a new PD Term Detail
@login_required
def delinquent_pd_create(request):
    if request.method == 'POST':
        form = PDTermStructureDtlForm(request.POST)
        if form.is_valid():
            try:
                # Save the form without committing to the database yet
                delinquent_pd = form.save(commit=False)
                # Set the created_by field to the currently logged-in user
                delinquent_pd.created_by = request.user
                # Save the object to the database
                delinquent_pd.save()
                 # Log the creation in the AuditTrail
                AuditTrail.objects.create(
                    user=request.user,
                    model_name='Ldn_PD_Term_Structure_Dtl',
                    action='create',
                    object_id=delinquent_pd.pk,
                    change_description=(
                        f"Created Delinquent PD Term: Structure ID - {delinquent_pd.v_pd_term_structure_id}, "
                        f"Date - {delinquent_pd.fic_mis_date}, Risk Basis - {delinquent_pd.v_credit_risk_basis_cd}, "
                        f"PD Percent - {delinquent_pd.n_pd_percent}"
                    ),
                    timestamp=now(),
                )

                messages.success(request, "Delinquent PD Term added successfully!")
                return redirect('delinquent_pd_list')
            except Exception as e:
                # Capture and display the specific exception message
                messages.error(request, f"Error adding Delinquent PD Term: {e}")
        else:
            # Display form validation errors
            messages.error(request, "There were validation errors. Please correct them below.")
    else:
        form = PDTermStructureDtlForm()
    
    return render(request, 'probability_conf/delinquent_pd_form.html', {'form': form})

# Edit PD Term Detail
@login_required
def delinquent_pd_edit(request, term_id):
    pd_term_detail = get_object_or_404(Ldn_PD_Term_Structure_Dtl, pk=term_id)
    
    if request.method == 'POST':
        form = PDTermStructureDtlForm(request.POST, instance=pd_term_detail)
        if form.is_valid():
            try:
                # Save the form without committing to the database yet
                delinquent_pd = form.save(commit=False)
                # Set the created_by field to the currently logged-in user
                delinquent_pd.created_by = request.user
                # Save the object to the database
                delinquent_pd.save()

                # Log the update in the AuditTrail
                AuditTrail.objects.create(
                    user=request.user,
                    model_name='Ldn_PD_Term_Structure_Dtl',
                    action='update',
                    object_id=delinquent_pd.pk,
                    change_description=(
                        f"Updated Delinquent PD Term: Structure ID - {delinquent_pd.v_pd_term_structure_id}, "
                        f"Date - {delinquent_pd.fic_mis_date}, Risk Basis - {delinquent_pd.v_credit_risk_basis_cd}, "
                        f"PD Percent - {delinquent_pd.n_pd_percent}"
                    ),
                    timestamp=now(),
                )


                messages.success(request, "Delinquent PD Term updated successfully!")
                return redirect('delinquent_pd_list')
            except Exception as e:
                # Capture and display the specific exception message
                messages.error(request, f"Error updating Delinquent PD Term: {e}")
        else:
            # Display form validation errors
            messages.error(request, "There were validation errors. Please correct them below.")
    else:
        form = PDTermStructureDtlForm(instance=pd_term_detail)
    
    return render(request, 'probability_conf/delinquent_pd_form.html', {'form': form})
# Delete PD Term Detail
@login_required
def delinquent_pd_delete(request, term_id):
    pd_term_detail = get_object_or_404(Ldn_PD_Term_Structure_Dtl, pk=term_id)
    if request.method == 'POST':
         # Log the deletion in the AuditTrail
        AuditTrail.objects.create(
            user=request.user,
            model_name='Ldn_PD_Term_Structure_Dtl',
            action='delete',
            object_id=pd_term_detail.pk,
            change_description=(
                f"Deleted Delinquent Based PD Term: Structure ID - {pd_term_detail.v_pd_term_structure_id}, "
                f"Date - {pd_term_detail.fic_mis_date}, Risk Basis - {pd_term_detail.v_credit_risk_basis_cd}, "
                f"PD Percent - {pd_term_detail.n_pd_percent}"
            ),
            timestamp=now(),
        )
        pd_term_detail.delete()
        messages.success(request, "Delinquent PD Term deleted successfully!")
        return redirect('delinquent_pd_list')
    return render(request, 'probability_conf/delinquent_pd_confirm_delete.html', {'pd_term_detail': pd_term_detail})

# List View
# List all Rating Based PD Terms
@login_required
def rating_pd_list(request):
    # Filter for 'Rating' type PD Term Structures
    pd_term_details_list = Ldn_PD_Term_Structure_Dtl.objects.filter(
        v_pd_term_structure_id__v_pd_term_structure_type='R'
    ).select_related('v_pd_term_structure_id')
    
    # Set up pagination with 5 items per page
    paginator = Paginator(pd_term_details_list, 5)
    page_number = request.GET.get('page')
    pd_term_details = paginator.get_page(page_number)

    return render(request, 'probability_conf/rating_pd_list.html', {'pd_term_details': pd_term_details})

# Create a new Rating Based PD Term Detail
@login_required
def rating_pd_create(request):
    if request.method == 'POST':
        form = PDTermStructureDtlRatingForm(request.POST)
        if form.is_valid():
            try:
                # Save the form without committing to the database yet
                rating = form.save(commit=False)
                # Set the created_by field to the currently logged-in user
                rating.created_by = request.user
                # Save the object to the database
                rating.save()
                # Log the creation in the AuditTrail
                AuditTrail.objects.create(
                    user=request.user,
                    model_name='Ldn_PD_Term_Structure_Dtl',
                    action='create',
                    object_id=rating.pk,
                    change_description=(
                        f"Created Rating Based PD Term: Structure ID - {rating.v_pd_term_structure_id}, "
                        f"Date - {rating.fic_mis_date}, Risk Basis - {rating.v_credit_risk_basis_cd}, "
                        f"PD Percent - {rating.n_pd_percent}"
                    ),
                    timestamp=now(),
                )

                messages.success(request, "Rating Based PD Term added successfully!")
                return redirect('rating_pd_list')
            except Exception as e:
                # Capture and display the specific exception message
                messages.error(request, f"Error adding Rating Based PD Term: {e}")
    else:
        form = PDTermStructureDtlRatingForm()
    
    return render(request, 'probability_conf/rating_pd_form.html', {'form': form})

# Edit Rating Based PD Term Detail
@login_required
def rating_pd_edit(request, term_id):
    pd_term_detail = get_object_or_404(Ldn_PD_Term_Structure_Dtl, pk=term_id)
    if request.method == 'POST':
        form = PDTermStructureDtlRatingForm(request.POST, instance=pd_term_detail)
        if form.is_valid():
            try:
                # Save the form without committing to the database yet
                rating = form.save(commit=False)
                # Set the created_by field to the currently logged-in user
                rating.created_by = request.user
                # Save the object to the database
                rating.save()

                AuditTrail.objects.create(
                    user=request.user,
                    model_name='Ldn_PD_Term_Structure_Dtl',
                    action='update',
                    object_id=rating.pk,
                    change_description=(
                        f"Updated Rating Based PD Term: Structure ID - {rating.v_pd_term_structure_id}, "
                        f"Date - {rating.fic_mis_date}, Risk Basis - {rating.v_credit_risk_basis_cd}, "
                        f"PD Percent - {rating.n_pd_percent}"
                    ),
                    timestamp=now(),
                )

                
                messages.success(request, "Rating Based PD Term updated successfully!")
                return redirect('rating_pd_list')
            except Exception as e:
                # Capture and display the specific exception message
                messages.error(request, f"Error updating Rating Based PD Term: {e}")
        
    else:
        form = PDTermStructureDtlRatingForm(instance=pd_term_detail)
    
    return render(request, 'probability_conf/rating_pd_form.html', {'form': form})

# Delete Rating Based PD Term Detail
@login_required
def rating_pd_delete(request, term_id):
    pd_term_detail = get_object_or_404(Ldn_PD_Term_Structure_Dtl, pk=term_id)
    if request.method == 'POST':
         # Log the deletion in the AuditTrail
        AuditTrail.objects.create(
            user=request.user,
            model_name='Ldn_PD_Term_Structure_Dtl',
            action='delete',
            object_id=pd_term_detail.pk,
            change_description=(
                f"Deleted Rating Based PD Term: Structure ID - {pd_term_detail.v_pd_term_structure_id}, "
                f"Date - {pd_term_detail.fic_mis_date}, Risk Basis - {pd_term_detail.v_credit_risk_basis_cd}, "
                f"PD Percent - {pd_term_detail.n_pd_percent}"
            ),
            timestamp=now(),
        )
        pd_term_detail.delete()
        messages.success(request, "Rating Based PD Term deleted successfully!")
        return redirect('rating_pd_list')
    return render(request, 'probability_conf/rating_pd_confirm_delete.html', {'pd_term_detail': pd_term_detail})



# List all Interpolation Methods
@login_required
def interpolation_method_list(request):
    preferences = FSI_LLFP_APP_PREFERENCES.objects.all()
    return render(request, 'probability_conf/interpolation_method_list.html', {'preferences': preferences})

# Create a new Interpolation Method
@login_required
def interpolation_method_create(request):
    if request.method == 'POST':
        form = InterpolationMethodForm(request.POST)
        if form.is_valid():
            try:
                # Save the form without committing to the database yet
                interpolation = form.save(commit=False)
                # Set the created_by field to the currently logged-in user
                interpolation.created_by = request.user
                # Save the object to the database
                interpolation.save()

                # Log the action to the AuditTrail
                action =  "create"
                AuditTrail.objects.create(
                    user=request.user,
                    model_name='FSI_LLFP_APP_PREFERENCES',
                    action=action,
                    object_id=interpolation.pk,
                    change_description=(
                        f"{action.title()}d Interpolation Method: "
                        f"Method - {interpolation.get_pd_interpolation_method_display()}, "
                        f"Level - {interpolation.get_interpolation_level_display()}"
                    ),
                    timestamp=now(),
                )

                messages.success(request, "Interpolation Method added successfully!")
                return redirect('interpolation_method_list')
            except Exception as e:
                # Capture and display the specific exception message
                messages.error(request, f"Error adding Interpolation Method: {e}")
        else:
            # Display form validation errors
            messages.error(request, "There were validation errors. Please correct them below.")
    else:
        form = InterpolationMethodForm()

    return render(request, 'probability_conf/interpolation_method_form.html', {'form': form})
# Edit Interpolation Method
@login_required
def interpolation_method_edit(request, method_id):
    interpolation_method = get_object_or_404(FSI_LLFP_APP_PREFERENCES, pk=method_id)
    
    if request.method == 'POST':
        form = InterpolationMethodForm(request.POST, instance=interpolation_method)
        if form.is_valid():
            try:
                # Save the form without committing to the database yet
                interpolation = form.save(commit=False)
                # Set the created_by field to the currently logged-in user
                interpolation.created_by = request.user
                # Save the object to the database
                interpolation.save()

                # Log the action to the AuditTrail
                action =  "update"
                AuditTrail.objects.create(
                    user=request.user,
                    model_name='FSI_LLFP_APP_PREFERENCES',
                    action=action,
                    object_id=interpolation.pk,
                    change_description=(
                        f"{action.title()}d Interpolation Method: "
                        f"Method - {interpolation.get_pd_interpolation_method_display()}, "
                        f"Level - {interpolation.get_interpolation_level_display()}"
                    ),
                    timestamp=now(),
                )
                
                messages.success(request, "Interpolation Method updated successfully!")
                return redirect('interpolation_method_list')
            except Exception as e:
                # Capture and display the specific exception message
                messages.error(request, f"Error updating Interpolation Method: {e}")
        else:
            # Display form validation errors
            messages.error(request, "There were validation errors. Please correct them below.")
    else:
        form = InterpolationMethodForm(instance=interpolation_method)

    return render(request, 'probability_conf/interpolation_method_form.html', {'form': form})

# Delete Interpolation Method
@login_required
def interpolation_method_delete(request, method_id):
    interpolation_method = get_object_or_404(FSI_LLFP_APP_PREFERENCES, pk=method_id)

    if request.method == 'POST':
        # Log the deletion to the AuditTrail before deletion
        AuditTrail.objects.create(
            user=request.user,
            model_name='FSI_LLFP_APP_PREFERENCES',
            action='delete',
            object_id=interpolation_method.pk,
            change_description=f"Deleted Interpolation Method: {interpolation_method}",
            timestamp=now(),
        )

        # Perform the actual deletion
        interpolation_method.delete()

        messages.success(request, "Interpolation Method deleted successfully!")
        return redirect('interpolation_method_list')

    return render(request, 'probability_conf/interpolation_method_confirm_delete.html', {'interpolation_method': interpolation_method})


@login_required
def reporting_home(request):
    return render(request, 'reports/reporting.html')

@login_required
def list_reports(request):
    # This view will render the list of available reports
    return render(request, 'reports/list_reports.html')




@login_required
@require_http_methods(["GET", "POST"])
def view_results_and_extract(request):
    # Handle AJAX request for dynamic Run Key loading (unchanged)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' and request.GET.get('field') == 'fic_mis_date':
        fic_mis_date = request.GET.get('value')
        n_run_keys = FCT_Reporting_Lines.objects.filter(fic_mis_date=fic_mis_date).order_by('-n_run_key').values_list('n_run_key', flat=True).distinct()
        return JsonResponse({'n_run_keys': list(n_run_keys) if n_run_keys.exists() else []})

    # Non-AJAX processing
    fic_mis_dates = FCT_Reporting_Lines.objects.order_by('-fic_mis_date').values_list('fic_mis_date', flat=True).distinct()
    fic_mis_date = request.GET.get('fic_mis_date')
    

    # Save selected filters to the session
    if fic_mis_date :
        request.session['fic_mis_date'] = fic_mis_date
       
    else:
        request.session.pop('fic_mis_date', None)
     

    # Ensure filters are provided
    if not fic_mis_date :
        messages.error(request, "Both Reporting Date and Run Key are required.")
        return render(request, 'reports/report_view.html', {
            'selected_columns': [],
            'report_data': [],
            'filters': request.GET,
            'fic_mis_dates': fic_mis_dates,
            'fic_mis_date': fic_mis_date,
       
        })

    # Apply filters and fetch data (unchanged)
    filters = {'fic_mis_date': fic_mis_date,}
    report_config = get_object_or_404(ReportColumnConfig, report_name="default_report")
    selected_columns = report_config.selected_columns
    report_data = FCT_Reporting_Lines.objects.filter(**filters).values(*selected_columns)

    paginator = Paginator(report_data, 25)
    page = request.GET.get('page', 1)
    try:
        paginated_report_data = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        paginated_report_data = paginator.page(1)

    return render(request, 'reports/report_view.html', {
        'selected_columns': selected_columns,
        'report_data': paginated_report_data,
        'filters': filters,
        'fic_mis_dates': fic_mis_dates,
        'fic_mis_date': fic_mis_date,
      
    })


@login_required
def download_report(request):
    # Retrieve filters from the session
    fic_mis_date = request.session.get('fic_mis_date')
   

    # Validate filters
    if not fic_mis_date or not n_run_key:
        return HttpResponse("Both Reporting Date and Run Key are required.", status=400)

    # Apply filters to fetch data
    filters = {'fic_mis_date': fic_mis_date,}
    try:
        report_config = ReportColumnConfig.objects.get(report_name="default_report")
        selected_columns = report_config.selected_columns
    except ReportColumnConfig.DoesNotExist:
        return HttpResponse("Report configuration not found.", status=404)

    report_data = FCT_Reporting_Lines.objects.filter(**filters).values(*selected_columns)
    if not report_data.exists():
        return HttpResponse("No data found for the selected filters.", status=404)

    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="filtered_report.csv"'
    writer = csv.writer(response)
    writer.writerow(selected_columns)

    for row in report_data:
        writer.writerow([row[column] for column in selected_columns])

    return response


############################################

# Directory to save the CSV files
CSV_DIR = os.path.join(os.getcwd(), 'csv_files')

@login_required
@require_http_methods(["GET", "POST"])
def ecl_main_filter_view(request):
    # Handle POST request (applying the filter)
    if request.method == 'POST':
        # Retrieve selected Reporting Date and Run Key from the form
        fic_mis_date = request.POST.get('fic_mis_date')
 
        
        if fic_mis_date :
            # Store the selected filter values in session for later use
            request.session['fic_mis_date'] = fic_mis_date
            

            # Retrieve filtered data based on the selected main filter values
            ecl_data = FCT_Reporting_Lines.objects.filter(fic_mis_date=fic_mis_date)

            # Convert the filtered data into a DataFrame
            ecl_data_df = pd.DataFrame(list(ecl_data.values()))

            # Convert date fields to strings
            if 'fic_mis_date' in ecl_data_df.columns:
                ecl_data_df['fic_mis_date'] = ecl_data_df['fic_mis_date'].astype(str)
            if 'd_maturity_date' in ecl_data_df.columns:
                ecl_data_df['d_maturity_date'] = ecl_data_df['d_maturity_date'].astype(str)

            # Create the directory if it doesn't exist
            if not os.path.exists(CSV_DIR):
                os.makedirs(CSV_DIR)

            # Save the data as a CSV file in the session (store the filename)
            csv_filename = os.path.join(CSV_DIR, f"ecl_data_{fic_mis_date}.csv")
            ecl_data_df.to_csv(csv_filename, index=False)
            request.session['csv_filename'] = csv_filename

            # Redirect to the sub-filter view
            return redirect('ecl_sub_filter_view')

    # Handle GET request
    fic_mis_dates = FCT_Reporting_Lines.objects.order_by('-fic_mis_date').values_list('fic_mis_date', flat=True).distinct()

    # Check for AJAX requests to dynamically update run key dropdown
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        field_name = request.GET.get('field')
        field_value = request.GET.get('value')

        if field_name == 'fic_mis_date':
            # Fetch run keys corresponding to the selected FIC MIS date
            n_run_keys = FCT_Reporting_Lines.objects.filter(fic_mis_date=field_value).order_by('-n_run_key').values_list('n_run_key', flat=True).distinct()
            return JsonResponse({'n_run_keys': list(n_run_keys)})

    # Render the main filter template with Reporting Dates
    return render(request, 'reports/ecl_summary_report.html', {'fic_mis_dates': fic_mis_dates})




@login_required
@require_http_methods(["GET", "POST"])
def ecl_sub_filter_view(request):
    # Retrieve the CSV filename from the session
    csv_filename = request.session.get('csv_filename')

    # If no data is available, redirect to the main filter page
    if not csv_filename or not os.path.exists(csv_filename):
        return redirect('ecl_main_filter_view')

    # Load the data from the CSV file
    ecl_data_df = pd.read_csv(csv_filename)

    # Retrieve sub-filter form fields from the request
    n_prod_segment = request.GET.get('n_prod_segment')
    n_prod_type = request.GET.get('n_prod_type')
    n_stage_descr = request.GET.get('n_stage_descr')
    n_loan_type = request.GET.get('n_loan_type')

    # Apply sub-filters if provided
    if n_prod_segment:
        ecl_data_df = ecl_data_df[ecl_data_df['n_prod_segment'] == n_prod_segment]
    if n_prod_type:
        ecl_data_df = ecl_data_df[ecl_data_df['n_prod_type'] == n_prod_type]
    if n_stage_descr:
        ecl_data_df = ecl_data_df[ecl_data_df['n_stage_descr'] == n_stage_descr]
    if n_loan_type:
        ecl_data_df = ecl_data_df[ecl_data_df['n_loan_type'] == n_loan_type]

    # Retrieve the selected group by field from the request (GET or POST)
    group_by_field = request.GET.get('group_by_field', 'n_stage_descr')  # Default group by 'n_stage_descr'

    # Group the data by the selected field and sum the amounts, while also counting unique accounts
    grouped_data = (
        ecl_data_df.groupby(group_by_field)
        .agg({
            'n_exposure_at_default_ncy': 'sum',
            'n_exposure_at_default_rcy': 'sum',
            'n_12m_ecl_rcy': 'sum',
            'n_lifetime_ecl_rcy': 'sum',
            'n_account_number': pd.Series.nunique,  # Count unique accounts in each group
        })
        .reset_index()
        .to_dict(orient='records')
    )

    # Calculate grand totals
    grand_totals = {
        'n_exposure_at_default_ncy': ecl_data_df['n_exposure_at_default_ncy'].sum(),
        'n_exposure_at_default_rcy': ecl_data_df['n_exposure_at_default_rcy'].sum(),
        'n_12m_ecl_rcy': ecl_data_df['n_12m_ecl_rcy'].sum(),
        'n_lifetime_ecl_rcy': ecl_data_df['n_lifetime_ecl_rcy'].sum(),
        'n_account_number': ecl_data_df['n_account_number'].nunique(),  # Grand total for unique accounts
    }

    # Calculate percentages for the second table
    grouped_data_percentages = []
    total_ecl_12m = grand_totals['n_12m_ecl_rcy']
    total_ecl_lifetime = grand_totals['n_lifetime_ecl_rcy']
    total_accounts = grand_totals['n_account_number']

    for row in grouped_data:
        grouped_data_percentages.append({
            group_by_field: row[group_by_field],
            'percent_12m_ecl_rcy': (row['n_12m_ecl_rcy'] / total_ecl_12m * 100) if total_ecl_12m else 0,
            'percent_lifetime_ecl_rcy': (row['n_lifetime_ecl_rcy'] / total_ecl_lifetime * 100) if total_ecl_lifetime else 0,
            'percent_accounts': (row['n_account_number'] / total_accounts * 100) if total_accounts else 0,
        })

    # Distinct values for sub-filters
    distinct_prod_segments = list(ecl_data_df['n_prod_segment'].unique())
    distinct_prod_types = list(ecl_data_df['n_prod_type'].unique())
    distinct_stage_descrs = list(ecl_data_df['n_stage_descr'].unique())
    distinct_loan_types = list(ecl_data_df['n_loan_type'].unique())

    # Store the grouped data and grand totals in the session for Excel export
    request.session['grouped_data'] = grouped_data
    request.session['group_by_field'] = group_by_field
    request.session['grand_totals'] = grand_totals
    request.session['grouped_data_percentages'] = grouped_data_percentages

    # Render the sub-filter view template
    return render(request, 'reports/ecl_summary_report_sub.html', {
        'grouped_data': grouped_data,
        'grouped_data_percentages': grouped_data_percentages,
        'group_by_field': group_by_field,
        'distinct_prod_segments': distinct_prod_segments,
        'distinct_prod_types': distinct_prod_types,
        'distinct_stage_descrs': distinct_stage_descrs,
        'distinct_loan_types': distinct_loan_types,
        'grand_totals': grand_totals,
    })


# Export to Excel dynamically based on the current filtered and grouped data
@login_required
def export_ecl_report_to_excel(request):
    # Retrieve the filtered data, grouped data, grouped percentages, and grand totals from session
    grouped_data = request.session.get('grouped_data', [])
    grouped_data_percentages = request.session.get('grouped_data_percentages', [])
    group_by_field = request.session.get('group_by_field', 'n_stage_descr')
    grand_totals = request.session.get('grand_totals', {})

    # Convert the grouped data into pandas DataFrames
    df_grouped = pd.DataFrame(grouped_data)
    df_percentages = pd.DataFrame(grouped_data_percentages)

    # Check if percentages data exists and add missing fields
    if not df_percentages.empty and group_by_field not in df_percentages.columns:
        df_percentages[group_by_field] = [row[group_by_field] for row in grouped_data]

    # Create a new Excel workbook
    wb = openpyxl.Workbook()

    # Add the first sheet for the absolute numbers
    ws1 = wb.active
    ws1.title = "ECL Report (Absolute)"
    headers = [group_by_field, "EAD Orig Currency", "EAD Reporting Currency", "12 Month Reporting ECL", "Lifetime Reporting ECL", "Number of Accounts"]
    ws1.append(headers)

    # Add the grouped data to the first sheet
    for row in dataframe_to_rows(df_grouped, index=False, header=False):
        ws1.append(row)

    # Add the grand totals at the bottom of the first sheet
    ws1.append([
        'Grand Total',
        grand_totals['n_exposure_at_default_ncy'],
        grand_totals['n_exposure_at_default_rcy'],
        grand_totals['n_12m_ecl_rcy'],
        grand_totals['n_lifetime_ecl_rcy'],
        grand_totals['n_account_number']
    ])

    # Add styling to the first sheet
    style_excel_sheet(ws1, len(grouped_data), headers)

    # Add a second sheet for percentages
    ws2 = wb.create_sheet(title="ECL Report (Percentages)")
    percentage_headers = [group_by_field, "% of 12 Month Reporting ECL", "% of Lifetime Reporting ECL", "% of Number of Accounts"]
    ws2.append(percentage_headers)

    # Ensure percentages data is not empty before adding rows
    if not df_percentages.empty:
        for row in dataframe_to_rows(df_percentages, index=False, header=False):
            ws2.append(row)

    # Add styling to the second sheet
    style_excel_sheet(ws2, len(grouped_data_percentages), percentage_headers)

    # Create an HTTP response with an Excel attachment
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ecl summary report.xlsx"'

    # Save the workbook to the response
    wb.save(response)
    return response

def style_excel_sheet(ws, data_row_count, headers):
    """Apply styling to an Excel sheet."""
    header_fill = PatternFill(start_color="2d5c8e", end_color="2d5c8e", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(horizontal="center", vertical="center")

    # Style the header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment

    # Apply zebra striping for data rows
    light_fill = PatternFill(start_color="d1e7dd", end_color="d1e7dd", fill_type="solid")
    for row in ws.iter_rows(min_row=2, max_row=data_row_count + 1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.fill = light_fill

    # Apply bold font for the grand total row if it exists
    if data_row_count > 0:
        for cell in ws[data_row_count + 1]:
            cell.font = Font(bold=True)

    # Adjust column widths
    for column in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value is not None)
        ws.column_dimensions[column[0].column_letter].width = max_length + 2




CSV_DIR = os.path.join(os.getcwd(), 'csv_files')
@login_required
@require_http_methods(["GET", "POST"])
def ecl_reconciliation_main_filter_view(request):
    # Initialize an empty errors dictionary
    errors = {}

    # Handle POST request (applying the filter)
    if request.method == 'POST':
        # Retrieve selected Reporting Dates and Run Keys from the form
        fic_mis_date1 = request.POST.get('fic_mis_date1')
      
        fic_mis_date2 = request.POST.get('fic_mis_date2')
       

        # Validate that both dates and run keys are provided
        if not fic_mis_date1:
            errors['fic_mis_date1'] = 'Please select Reporting Date 1.'
   
        if not fic_mis_date2:
            errors['fic_mis_date2'] = 'Please select Reporting Date 2.'
    

        # If there are no errors, proceed to filtering
        if not errors:
            # Store the selected filter values in session for later use
            request.session['fic_mis_date1'] = fic_mis_date1
        
            request.session['fic_mis_date2'] = fic_mis_date2
      

            # Retrieve filtered data based on the selected main filter values
            ecl_data1 = FCT_Reporting_Lines.objects.filter(fic_mis_date=fic_mis_date1)
            ecl_data2 = FCT_Reporting_Lines.objects.filter(fic_mis_date=fic_mis_date2)

            # Convert the filtered data into a DataFrame
            ecl_data1_df = pd.DataFrame(list(ecl_data1.values()))
            ecl_data2_df = pd.DataFrame(list(ecl_data2.values()))

            # Convert date fields to strings for both DataFrames
            for df in [ecl_data1_df, ecl_data2_df]:
                if 'fic_mis_date' in df.columns:
                    df['fic_mis_date'] = df['fic_mis_date'].astype(str)
                if 'd_maturity_date' in df.columns:
                    df['d_maturity_date'] = df['d_maturity_date'].astype(str)

            # Create the directory if it doesn't exist
            if not os.path.exists(CSV_DIR):
                os.makedirs(CSV_DIR)

            # Save the data as CSV files in the session (store the filenames)
            csv_filename1 = os.path.join(CSV_DIR, f"ecl_data_{fic_mis_date1}.csv")
            csv_filename2 = os.path.join(CSV_DIR, f"ecl_data_{fic_mis_date2}.csv")
            
            ecl_data1_df.to_csv(csv_filename1, index=False)
            ecl_data2_df.to_csv(csv_filename2, index=False)
            # Store the filenames in the session and explicitly mark the session as modified
            request.session['csv_filename1'] = csv_filename1
            request.session['csv_filename2'] = csv_filename2
            request.session.modified = True
            # Redirect to the sub-filter view
            return redirect('ecl_reconciliation_sub_filter_view')

    # Handle GET request to load Reporting Dates
    fic_mis_dates = FCT_Reporting_Lines.objects.order_by('-fic_mis_date').values_list('fic_mis_date', flat=True).distinct()

    # Check for AJAX requests to dynamically update run key dropdown
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        field_name = request.GET.get('field')
        field_value = request.GET.get('value')

        if field_name == 'fic_mis_date':
            # Fetch run keys corresponding to the selected FIC MIS date
            n_run_keys = FCT_Reporting_Lines.objects.filter(fic_mis_date=field_value).order_by('-n_run_key').values_list('n_run_key', flat=True).distinct()
            return JsonResponse({'n_run_keys': list(n_run_keys)})

    # If there are errors or it's a GET request, render the form with any errors
    return render(request, 'reports/ecl_reconciliation_main_filter.html', {
        'fic_mis_dates': fic_mis_dates,
        'errors': errors  # Pass the errors to the template
    })



@login_required
@require_http_methods(["GET", "POST"])
def ecl_reconciliation_sub_filter_view(request):
    # Retrieve the CSV filenames from the session
    csv_filename_1 = request.session.get('csv_filename1')
    csv_filename_2 = request.session.get('csv_filename2')

    # Check if the CSV files exist and display error if missing
    errors = []
    if not csv_filename_1:
        errors.append('The first dataset is missing. Please select a valid Reporting Date 1 and Run Key 1.')
    elif not os.path.exists(csv_filename_1):
        errors.append(f"The file for Reporting Date 1 does not exist: {csv_filename_1}")
    
    if not csv_filename_2:
        errors.append('The second dataset is missing. Please select a valid Reporting Date 2 and Run Key 2.')
    elif not os.path.exists(csv_filename_2):
        errors.append(f"The file for Reporting Date 2 does not exist: {csv_filename_2}")

    # If there are any errors, display them and do not proceed further
    if errors:
        for error in errors:
            messages.error(request, error)
        return render(request, 'reports/ecl_reconciliation_report_sub.html', {'errors': errors})

    # Load the data from the CSV files
    ecl_data_df_1 = pd.read_csv(csv_filename_1)
    ecl_data_df_2 = pd.read_csv(csv_filename_2)

    # Determine which dataframe has the higher and lower date based on fic_mis_date
    if ecl_data_df_1['fic_mis_date'].max() > ecl_data_df_2['fic_mis_date'].max():
        df_higher_date = ecl_data_df_1
        df_lower_date = ecl_data_df_2
    else:
        df_higher_date = ecl_data_df_2
        df_lower_date = ecl_data_df_1

    # Ensure n_prod_segment is the same type in both dataframes
    df_higher_date['n_prod_segment'] = df_higher_date['n_prod_segment'].astype(str)
    df_lower_date['n_prod_segment'] = df_lower_date['n_prod_segment'].astype(str)

    # Merge based on fic_mis_date and other keys
    merged_data = pd.merge(
        df_higher_date, 
        df_lower_date, 
        on=['fic_mis_date', 'n_run_key', 'n_stage_descr','v_ccy_code','n_prod_segment','n_prod_type','n_loan_type'],  
        how='outer', 
        suffixes=('_higher', '_lower')
    )

    # Fill NaN values with zeros for numeric fields
    merged_data.fillna({
        'n_exposure_at_default_ncy_higher': 0, 'n_exposure_at_default_ncy_lower': 0,
        'n_exposure_at_default_rcy_higher': 0, 'n_exposure_at_default_rcy_lower': 0,
        'n_12m_ecl_rcy_higher': 0, 'n_12m_ecl_rcy_lower': 0,
        'n_lifetime_ecl_rcy_higher': 0, 'n_lifetime_ecl_rcy_lower': 0
    }, inplace=True)

    # Apply sub-filters to the merged data
    v_ccy_code = request.GET.get('v_ccy_code')
    n_prod_segment = request.GET.get('n_prod_segment')
    n_prod_type = request.GET.get('n_prod_type')
    n_stage_descr = request.GET.get('n_stage_descr')
    n_loan_type = request.GET.get('n_loan_type')

    if v_ccy_code:
        merged_data = merged_data[merged_data['v_ccy_code'] == v_ccy_code]
    if n_prod_segment:
        merged_data = merged_data[merged_data['n_prod_segment'] == n_prod_segment]
    if n_prod_type:
        merged_data = merged_data[merged_data['n_prod_type'] == n_prod_type]
    if n_stage_descr:
        merged_data = merged_data[merged_data['n_stage_descr'] == n_stage_descr]
    if n_loan_type:
        merged_data = merged_data[merged_data['n_loan_type'] == n_loan_type]

    # Group by functionality after merging and applying filters
    group_by_field = request.GET.get('group_by_field', 'n_stage_descr') 

    # Group data and calculate the sum, while counting unique account numbers for higher and lower datasets
    grouped_data = merged_data.groupby(group_by_field).agg({
        'n_exposure_at_default_ncy_higher': 'sum',
        'n_exposure_at_default_ncy_lower': 'sum',
        'n_exposure_at_default_rcy_higher': 'sum',
        'n_exposure_at_default_rcy_lower': 'sum',
        'n_12m_ecl_rcy_higher': 'sum',
        'n_12m_ecl_rcy_lower': 'sum',
        'n_lifetime_ecl_rcy_higher': 'sum',
        'n_lifetime_ecl_rcy_lower': 'sum',
        # Counting distinct account numbers for higher and lower
        'n_account_number_higher': pd.Series.nunique,  
        'n_account_number_lower': pd.Series.nunique    
    }).reset_index()

    # Create new columns for account counts based on the 'n_account_number' column
    grouped_data['n_accounts_in_higher'] = grouped_data['n_account_number_higher']
    grouped_data['n_accounts_in_lower'] = grouped_data['n_account_number_lower']

    # Calculate differences (higher date minus lower date)
    grouped_data['difference_ead_ncy'] = grouped_data['n_exposure_at_default_ncy_higher'] - grouped_data['n_exposure_at_default_ncy_lower']
    grouped_data['difference_ead_rcy'] = grouped_data['n_exposure_at_default_rcy_higher'] - grouped_data['n_exposure_at_default_rcy_lower']
    grouped_data['difference_12m_ecl'] = grouped_data['n_12m_ecl_rcy_higher'] - grouped_data['n_12m_ecl_rcy_lower']
    grouped_data['difference_lifetime_ecl'] = grouped_data['n_lifetime_ecl_rcy_higher'] - grouped_data['n_lifetime_ecl_rcy_lower']

    # Status calculation based on 12-month ECL differences
    grouped_data['status_ead_ncy'] = grouped_data['difference_12m_ecl'].apply(
        lambda x: 'Increased' if x > 0 else ('Decreased' if x < 0 else 'No Change')
    )

    # Convert numeric columns to Python native types for JSON serialization
    grouped_data = grouped_data.applymap(lambda x: int(x) if isinstance(x, np.int64) else x)
    grouped_data = grouped_data.applymap(lambda x: float(x) if isinstance(x, np.float64) else x)

    # Convert to JSON-serializable format (a list of dictionaries) to store in session
    grouped_data_json = grouped_data.to_dict(orient='records')

    # Grand totals for columns and differences, also converted to Python native types
    grand_totals = {
        'n_exposure_at_default_ncy_higher': int(grouped_data['n_exposure_at_default_ncy_higher'].sum()),
        'n_exposure_at_default_rcy_higher': int(grouped_data['n_exposure_at_default_rcy_higher'].sum()),
        'n_12m_ecl_rcy_higher': int(grouped_data['n_12m_ecl_rcy_higher'].sum()),
        'n_lifetime_ecl_rcy_higher': int(grouped_data['n_lifetime_ecl_rcy_higher'].sum()),
        'n_exposure_at_default_ncy_lower': int(grouped_data['n_exposure_at_default_ncy_lower'].sum()),
        'n_exposure_at_default_rcy_lower': int(grouped_data['n_exposure_at_default_rcy_lower'].sum()),
        'n_12m_ecl_rcy_lower': int(grouped_data['n_12m_ecl_rcy_lower'].sum()),
        'n_lifetime_ecl_rcy_lower': int(grouped_data['n_lifetime_ecl_rcy_lower'].sum()),
        'difference_ead_rcy': int(grouped_data['difference_ead_rcy'].sum()),
        'difference_12m_ecl': int(grouped_data['difference_12m_ecl'].sum()),
        'difference_lifetime_ecl': int(grouped_data['difference_lifetime_ecl'].sum()),
        'n_accounts_in_higher': int(grouped_data['n_accounts_in_higher'].sum()),
        'n_accounts_in_lower': int(grouped_data['n_accounts_in_lower'].sum()),
    }
    # Calculate percentages for 12 Month ECL, Lifetime ECL, differences, and total accounts
    grouped_data['percent_12m_ecl_higher'] = (grouped_data['n_12m_ecl_rcy_higher'] / grand_totals['n_12m_ecl_rcy_higher'] * 100) if grand_totals['n_12m_ecl_rcy_higher'] else 0
    grouped_data['percent_12m_ecl_lower'] = (grouped_data['n_12m_ecl_rcy_lower'] / grand_totals['n_12m_ecl_rcy_lower'] * 100) if grand_totals['n_12m_ecl_rcy_lower'] else 0
    grouped_data['percent_lifetime_ecl_higher'] = (grouped_data['n_lifetime_ecl_rcy_higher'] / grand_totals['n_lifetime_ecl_rcy_higher'] * 100) if grand_totals['n_lifetime_ecl_rcy_higher'] else 0
    grouped_data['percent_lifetime_ecl_lower'] = (grouped_data['n_lifetime_ecl_rcy_lower'] / grand_totals['n_lifetime_ecl_rcy_lower'] * 100) if grand_totals['n_lifetime_ecl_rcy_lower'] else 0
    grouped_data['percent_total_accounts_higher'] = (grouped_data['n_accounts_in_higher'] / grand_totals['n_accounts_in_higher'] * 100) if grand_totals['n_accounts_in_higher'] else 0
    grouped_data['percent_total_accounts_lower'] = (grouped_data['n_accounts_in_lower'] / grand_totals['n_accounts_in_lower'] * 100) if grand_totals['n_accounts_in_lower'] else 0
    # Calculate percentage differences based on higher values
    grouped_data['percent_difference_12m_ecl'] = (grouped_data['difference_12m_ecl'] / grouped_data['n_12m_ecl_rcy_higher'] * 100).replace([np.inf, -np.inf], 0).fillna(0)
    grouped_data['percent_difference_lifetime_ecl'] = (grouped_data['difference_lifetime_ecl'] / grouped_data['n_lifetime_ecl_rcy_higher'] * 100).replace([np.inf, -np.inf], 0).fillna(0)
    grouped_data['percent_difference_ead_rcy'] = (grouped_data['difference_ead_rcy'] / grouped_data['n_exposure_at_default_rcy_higher'] * 100).replace([np.inf, -np.inf], 0).fillna(0)


    # Convert grouped_data to JSON-serializable format for percentages
    grouped_data_percentages = grouped_data.to_dict(orient='records')

    # Store the grouped data and grand totals in the session for Excel export (JSON serializable format)
    request.session['grouped_data'] = grouped_data_json
    request.session['group_by_field'] = group_by_field
    request.session['grand_totals'] = grand_totals
    request.session['grouped_data_percentages'] = grouped_data_percentages

    # Distinct values for sub-filters
    distinct_currency_codes = list(merged_data['v_ccy_code'].unique())
    distinct_prod_segments = list(merged_data['n_prod_segment'].unique())
    distinct_prod_types = list(merged_data['n_prod_type'].unique())
    distinct_stage_descrs = list(merged_data['n_stage_descr'].unique())
    distinct_loan_types = list(merged_data['n_loan_type'].unique())

    # Render the sub-filter view with the reconciliation data
    return render(request, 'reports/ecl_reconciliation_report_sub.html', {
        'grouped_data': grouped_data_json,
        'grouped_data_percentages': grouped_data_percentages,
        'group_by_field': group_by_field,
        'distinct_currency_codes': distinct_currency_codes,
        'distinct_prod_segments': distinct_prod_segments,
        'distinct_prod_types': distinct_prod_types,
        'distinct_stage_descrs': distinct_stage_descrs,
        'distinct_loan_types': distinct_loan_types,
        'grand_totals': grand_totals,
    })


@login_required
@require_http_methods(["POST"])
def export_ecl_reconciliation_to_excel(request):
    # Retrieve the filtered data, grouped percentages, and grand totals from session
    grouped_data = request.session.get('grouped_data', [])
    grouped_data_percentages = request.session.get('grouped_data_percentages', [])
    group_by_field = request.session.get('group_by_field', 'n_stage_descr')
    grand_totals = request.session.get('grand_totals', {})
    fic_mis_date1 = request.session.get('fic_mis_date1', '')
  
    fic_mis_date2 = request.session.get('fic_mis_date2', '')
   

    # Create a new Excel workbook
    wb = Workbook()

    # First Sheet: ECL Reconciliation Report (Absolute Values)
    ws1 = wb.active
    ws1.title = "ECL Reconciliation Report"

    # Merge cells for the headers
    ws1.merge_cells('B1:E1')  # Merge columns B to E in the first row for Date 1
    ws1['B1'] = f"Reporting Date 1 ({fic_mis_date1})"
    ws1['B1'].alignment = Alignment(horizontal='center', vertical='center')

    ws1.merge_cells('F1:I1')  # Merge columns F to I in the first row for Date 2
    ws1['F1'] = f"Reporting Date 2 ({fic_mis_date2}) "
    ws1['F1'].alignment = Alignment(horizontal='center', vertical='center')

    ws1.merge_cells('J1:L1')  # Merge columns J to L in the first row for Differences
    ws1['J1'] = "Differences"
    ws1['J1'].alignment = Alignment(horizontal='center', vertical='center')

    ws1.merge_cells('M1:N1')  # Merge columns M to N in the first row for Total Accounts
    ws1['M1'] = "Total Accounts"
    ws1['M1'].alignment = Alignment(horizontal='center', vertical='center')


    # Add the main headers
    headers = [
        group_by_field,
        f"Reporting Date 1 ({fic_mis_date1}) ",
        "",
        "",
        "",
        f"Reporting Date 2 ({fic_mis_date2})",
        "",
        "",
        "",
        "Differences",
        "",
        "",
        "Total Accounts",
        ""
    ]

    # Add sub-headers for the columns under the merged headers
    sub_headers = [
        group_by_field,
        "EAD Orig Currency",
        "EAD Reporting Currency",
        "12 Month ECL",
        "Lifetime ECL",
        "EAD Orig Currency",
        "EAD Reporting Currency",
        "12 Month ECL",
        "Lifetime ECL",
        "EAD Reporting Currency Difference",
        "12 Month ECL Difference",
        "Lifetime ECL Difference",
        f"Total Accounts ({fic_mis_date1})",
        f"Total Accounts ({fic_mis_date2})"
    ]
    ws1.append(sub_headers)

    # Add the grouped data to the first sheet
    for row in grouped_data:
        ws1.append([
            row.get(group_by_field, ''),
            row.get('n_exposure_at_default_ncy_higher', 0),
            row.get('n_exposure_at_default_rcy_higher', 0),
            row.get('n_12m_ecl_rcy_higher', 0),
            row.get('n_lifetime_ecl_rcy_higher', 0),
            row.get('n_exposure_at_default_ncy_lower', 0),
            row.get('n_exposure_at_default_rcy_lower', 0),
            row.get('n_12m_ecl_rcy_lower', 0),
            row.get('n_lifetime_ecl_rcy_lower', 0),
            row.get('difference_ead_rcy', 0),
            row.get('difference_12m_ecl', 0),
            row.get('difference_lifetime_ecl', 0),
            row.get('n_accounts_in_higher', 0),
            row.get('n_accounts_in_lower', 0),
        ])

    # Add the grand totals to the first sheet
    ws1.append([
        'Grand Total',
        grand_totals.get('n_exposure_at_default_ncy_higher', 0),
        grand_totals.get('n_exposure_at_default_rcy_higher', 0),
        grand_totals.get('n_12m_ecl_rcy_higher', 0),
        grand_totals.get('n_lifetime_ecl_rcy_higher', 0),
        grand_totals.get('n_exposure_at_default_ncy_lower', 0),
        grand_totals.get('n_exposure_at_default_rcy_lower', 0),
        grand_totals.get('n_12m_ecl_rcy_lower', 0),
        grand_totals.get('n_lifetime_ecl_rcy_lower', 0),
        grand_totals.get('difference_ead_rcy', 0),
        grand_totals.get('difference_12m_ecl', 0),
        grand_totals.get('difference_lifetime_ecl', 0),
        grand_totals.get('n_accounts_in_higher', 0),
        grand_totals.get('n_accounts_in_lower', 0),
    ])

    # Second Sheet: Percentage Table
    ws2 = wb.create_sheet(title="ECL Percentages")

     # Merge cells for the headers
    ws2.merge_cells('B1:C1')  # Merge columns B to E in the first row for Date 1
    ws2['B1'] = f"Reporting Date 1 ({fic_mis_date1}) "
    ws2['B1'].alignment = Alignment(horizontal='center', vertical='center')

    ws2.merge_cells('D1:E1')  # Merge columns F to I in the first row for Date 2
    ws2['D1'] = f"Reporting Date 2 ({fic_mis_date2})"
    ws2['D1'].alignment = Alignment(horizontal='center', vertical='center')

    ws2.merge_cells('F1:G1')  # Merge columns J to L in the first row for Differences
    ws2['F1'] = "Total Accounts"
    ws2['F1'].alignment = Alignment(horizontal='center', vertical='center')

    ws2.merge_cells('H1:J1')  # Merge columns M to N in the first row for Total Accounts
    ws2['H1'] = "Differences"
    ws2['H1'].alignment = Alignment(horizontal='center', vertical='center')

    # Add headers for percentages
    percentage_headers = [
        group_by_field,
        "% of 12 Month ECL (Date 1)",
        "% of Lifetime ECL (Date 1)",
        "% of 12 Month ECL (Date 2)",
        "% of Lifetime ECL (Date 2)",
        "% of Total Accounts (Date 1)",
        "% of Total Accounts (Date 2)",
        "% of EAD Reporting Currency Difference",
        "% of 12 Month ECL Difference",
        "% of Lifetime ECL Difference",
    ]
    ws2.append(percentage_headers)

    # Add grouped percentage data to the second sheet
    for row in grouped_data_percentages:
        ws2.append([
            row.get(group_by_field, ''),
            row.get('percent_12m_ecl_higher', 0),
            row.get('percent_lifetime_ecl_higher', 0),
            row.get('percent_12m_ecl_lower', 0),
            row.get('percent_lifetime_ecl_lower', 0),
            row.get('percent_total_accounts_higher', 0),
            row.get('percent_total_accounts_lower', 0),
            row.get('percent_difference_ead_rcy', 0),
            row.get('percent_difference_12m_ecl', 0),
            row.get('percent_difference_lifetime_ecl', 0),
        ])

    # Apply styling to both sheets (optional: customize as needed)
    style_excel_sheet(ws1, len(grouped_data), sub_headers)
    style_excel_sheet(ws2, len(grouped_data_percentages), percentage_headers)

    # Create an HTTP response with an Excel attachment
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="ecl_reconciliation_{fic_mis_date1}_{fic_mis_date2}.xlsx"'

    # Save the workbook to the response
    wb.save(response)
    return response


def style_excel_sheet(ws, data_row_count, headers):
    """Applies consistent styling to the Excel sheet."""
    header_fill = PatternFill(start_color="2d5c8e", end_color="2d5c8e", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(horizontal="center", vertical="center")

    # Style the header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment

    # Style the sub-headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment

    # Zebra striping for data rows
    light_fill = PatternFill(start_color="d1e7dd", end_color="d1e7dd", fill_type="solid")
    for row in ws.iter_rows(min_row=2, max_row=data_row_count + 3, min_col=1, max_col=len(headers)):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
            cell.fill = light_fill

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter if not isinstance(col[0], openpyxl.cell.cell.MergedCell) else None
        if column_letter:  # Only adjust if column is valid (not a merged cell)
            for cell in col:
                if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):  # Skip merged cells
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

CSV_DIR = os.path.join(os.getcwd(), 'csv_files')



@login_required
@require_http_methods(["GET", "POST"])
def water_fall_main_filter_view(request):
    errors = {}

    # Handle POST request (applying the filter)
    if request.method == 'POST':
        # Retrieve selected Reporting Dates and Run Keys from the form
        fic_mis_date1 = request.POST.get('fic_mis_date1')
      
        fic_mis_date2 = request.POST.get('fic_mis_date2')
       

        # Validate that both dates and run keys are provided
        if not fic_mis_date1:
            errors['fic_mis_date1'] = 'Please select Reporting Date 1.'
     
        if not fic_mis_date2:
            errors['fic_mis_date2'] = 'Please select Reporting Date 2.'
   

        # If no errors, proceed with filter
        if not errors:
            request.session['fic_mis_date1'] = fic_mis_date1
            request.session['fic_mis_date2'] = fic_mis_date2

            # Retrieve filtered data based on the selected main filter values
            ecl_data1 = FCT_Reporting_Lines.objects.filter(fic_mis_date=fic_mis_date1)
            ecl_data2 = FCT_Reporting_Lines.objects.filter(fic_mis_date=fic_mis_date2)

            # Convert the filtered data into a DataFrame
            ecl_data1_df = pd.DataFrame(list(ecl_data1.values()))
            ecl_data2_df = pd.DataFrame(list(ecl_data2.values()))

            # Ensure both DataFrames use the same set of account numbers
            common_account_numbers = set(ecl_data1_df['n_account_number']).intersection(ecl_data2_df['n_account_number'])
            ecl_data1_df = ecl_data1_df[ecl_data1_df['n_account_number'].isin(common_account_numbers)]
            ecl_data2_df = ecl_data2_df[ecl_data2_df['n_account_number'].isin(common_account_numbers)]

            # Convert date fields to strings for both DataFrames
            for df in [ecl_data1_df, ecl_data2_df]:
                if 'fic_mis_date' in df.columns:
                    df['fic_mis_date'] = df['fic_mis_date'].astype(str)
                if 'd_maturity_date' in df.columns:
                    df['d_maturity_date'] = df['d_maturity_date'].astype(str)

            # Create the directory if it doesn't exist
            if not os.path.exists(CSV_DIR):
                os.makedirs(CSV_DIR)

            # Save the data as CSV files in the session
            csv_filename1 = os.path.join(CSV_DIR, f"ecl_data_{fic_mis_date1}.csv")
            csv_filename2 = os.path.join(CSV_DIR, f"ecl_data_{fic_mis_date2}.csv")
            ecl_data1_df.to_csv(csv_filename1, index=False)
            ecl_data2_df.to_csv(csv_filename2, index=False)
            request.session['csv_filename1'] = csv_filename1
            request.session['csv_filename2'] = csv_filename2
            request.session.modified = True

            # Redirect to the sub-filter view
            return redirect('ecl_water_fall_sub_filter')

    # Handle GET request to load Reporting Dates
    fic_mis_dates = FCT_Reporting_Lines.objects.order_by('-fic_mis_date').values_list('fic_mis_date', flat=True).distinct()

    # AJAX request for dynamically updating the Run Key dropdowns
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        field_name = request.GET.get('field')
        field_value = request.GET.get('value')

        if field_name == 'fic_mis_date':
            # Fetch run keys corresponding to the selected FIC MIS date
            n_run_keys = FCT_Reporting_Lines.objects.filter(fic_mis_date=field_value).order_by('-n_run_key').values_list('n_run_key', flat=True).distinct()
            return JsonResponse({'n_run_keys': list(n_run_keys)})

    # If there are errors or it's a GET request, render the form with any errors
    return render(request, 'reports/waterfall_main_filter.html', {
        'fic_mis_dates': fic_mis_dates,
        'errors': errors  # Pass the errors to the template
    })


@login_required
@require_http_methods(["GET", "POST"])
def water_fall_sub_filter_view(request):
    # Retrieve CSV filenames from the session
    csv_filename_1 = request.session.get('csv_filename1')
    csv_filename_2 = request.session.get('csv_filename2')

    # Load CSV files into DataFrames
    ecl_data_df_1 = pd.read_csv(csv_filename_1)
    ecl_data_df_2 = pd.read_csv(csv_filename_2)

    # Merge data based on account number using an inner join
    merged_data = pd.merge(
        ecl_data_df_1,
        ecl_data_df_2,
        on=['n_account_number'],
        how='inner',
        suffixes=('_prev', '_curr')
    )

    # Filter based on account number from the request
    account_number = request.GET.get('n_account_number')
    if account_number:
        merged_data = merged_data[merged_data['n_account_number'] == account_number]

    # Prepare data for the waterfall report
    account_details = {}
    waterfall_data_12m = []
    waterfall_data_lifetime = []
    if not merged_data.empty:
        # Retrieve ECL values and other parameters for both periods
        twelve_month_ecl_prev = merged_data['n_12m_ecl_rcy_prev'].iloc[0]
        twelve_month_ecl_curr = merged_data['n_12m_ecl_rcy_curr'].iloc[0]
        lifetime_ecl_prev = merged_data['n_lifetime_ecl_rcy_prev'].iloc[0]
        lifetime_ecl_curr = merged_data['n_lifetime_ecl_rcy_curr'].iloc[0]

        # Calculate changes for the 12-month ECL waterfall
        stage_change_12m = twelve_month_ecl_curr - twelve_month_ecl_prev if merged_data['n_stage_descr_prev'].iloc[0] != merged_data['n_stage_descr_curr'].iloc[0] else 0
        exposure_change_12m = merged_data['n_exposure_at_default_rcy_curr'].iloc[0] - merged_data['n_exposure_at_default_rcy_prev'].iloc[0]
        pd_change_12m = merged_data['n_twelve_months_pd_curr'].iloc[0] - merged_data['n_twelve_months_pd_prev'].iloc[0]
        lgd_change_12m = merged_data['n_lgd_percent_curr'].iloc[0] - merged_data['n_lgd_percent_prev'].iloc[0]

        # Populate waterfall data for 12-month ECL
        waterfall_data_12m = [
            {"description": "Beginning 12-Month ECL", "impact": twelve_month_ecl_prev},
            {"description": "Change due to Stage (12-Month)", "impact": stage_change_12m},
            {"description": "Change in Exposure at Default (12-Month)", "impact": exposure_change_12m},
            {"description": "Change due to PD (12-Month)", "impact": pd_change_12m},
            {"description": "Change due to LGD (12-Month)", "impact": lgd_change_12m},
            {"description": "Ending 12-Month ECL", "impact": twelve_month_ecl_curr},
        ]

        # Calculate changes for the lifetime ECL waterfall
        stage_change_lifetime = lifetime_ecl_curr - lifetime_ecl_prev if merged_data['n_stage_descr_prev'].iloc[0] != merged_data['n_stage_descr_curr'].iloc[0] else 0
        exposure_change_lifetime = merged_data['n_exposure_at_default_rcy_curr'].iloc[0] - merged_data['n_exposure_at_default_rcy_prev'].iloc[0]
        pd_change_lifetime = merged_data['n_lifetime_pd_curr'].iloc[0] - merged_data['n_lifetime_pd_prev'].iloc[0]
        lgd_change_lifetime = merged_data['n_lgd_percent_curr'].iloc[0] - merged_data['n_lgd_percent_prev'].iloc[0]

        # Populate waterfall data for lifetime ECL
        waterfall_data_lifetime = [
            {"description": "Beginning Lifetime ECL", "impact": lifetime_ecl_prev},
            {"description": "Change due to Stage (Lifetime)", "impact": stage_change_lifetime},
            {"description": "Change in Exposure at Default (Lifetime)", "impact": exposure_change_lifetime},
            {"description": "Change due to PD (Lifetime)", "impact": pd_change_lifetime},
            {"description": "Change due to LGD (Lifetime)", "impact": lgd_change_lifetime},
            {"description": "Ending Lifetime ECL", "impact": lifetime_ecl_curr},
        ]

        # Account details for the selected account
        account_details = {
            'date_prev': merged_data['fic_mis_date_prev'].iloc[0],
            'date_curr': merged_data['fic_mis_date_curr'].iloc[0],
            'n_account_number': merged_data['n_account_number'].iloc[0],
            'exposure_at_default_prev': merged_data['n_exposure_at_default_rcy_prev'].iloc[0],
            'exposure_at_default_curr': merged_data['n_exposure_at_default_rcy_curr'].iloc[0],
            'ifrs_stage_prev': merged_data['n_stage_descr_prev'].iloc[0],
            'ifrs_stage_curr': merged_data['n_stage_descr_curr'].iloc[0],
            'twelve_month_pd_prev': merged_data['n_twelve_months_pd_prev'].iloc[0],
            'twelve_month_pd_curr': merged_data['n_twelve_months_pd_curr'].iloc[0],
            'lifetime_pd_prev': merged_data['n_lifetime_pd_prev'].iloc[0],
            'lifetime_pd_curr': merged_data['n_lifetime_pd_curr'].iloc[0],
            'lgd_prev': merged_data['n_lgd_percent_prev'].iloc[0],
            'lgd_curr': merged_data['n_lgd_percent_curr'].iloc[0],
            'twelve_month_ecl_prev': twelve_month_ecl_prev,
            'twelve_month_ecl_curr': twelve_month_ecl_curr,
            'lifetime_ecl_prev': lifetime_ecl_prev,
            'lifetime_ecl_curr': lifetime_ecl_curr,
            'prod_segment_prev': merged_data['n_prod_segment_prev'].iloc[0],
            'prod_segment_curr': merged_data['n_prod_segment_curr'].iloc[0],
        }

    # Get account options for the dropdown
    account_options = merged_data['n_account_number'].unique().tolist()

    return render(request, 'reports/waterfall_report.html', {
        'account_details': account_details,
        'account_options': account_options,
        'selected_account': account_number,
        'waterfall_data_12m': waterfall_data_12m,
        'waterfall_data_lifetime': waterfall_data_lifetime,
    })




@login_required
def configure_stages(request):
    title = "Stage Determination and Classification Configurations"
    return render(request, 'staging/staging_options.html', {'title': title})

@login_required
def staging_using_ratings(request):
    title = "Staging Using Ratings"
    # Additional logic for staging using ratings
    return render(request, 'staging/staging_using_ratings.html', {'title': title})

@login_required
def staging_using_delinquent_days(request):
    title = "Staging Using Delinquent Days"
    # Additional logic for staging using delinquent days
    return render(request, 'staging/staging_using_delinquent_days.html', {'title': title})


def stage_reassignment(request):
    title = "Stage Reassignment"
    # Additional logic for stage reassignment can go here
    return render(request, 'staging/stage_reassignment.html', {'title': title})


# List View for all credit ratings
class CreditRatingStageListView(LoginRequiredMixin,ListView):
    model = FSI_CreditRating_Stage
    template_name = 'staging/staging_using_ratings.html'  # Points to your template
    context_object_name = 'ratings'  # This will be used in the template
    paginate_by = 10  # You can adjust or remove pagination if not needed

    def get_queryset(self):
        queryset = FSI_CreditRating_Stage.objects.all()
        return queryset


# Create View for adding a new credit rating
class CreditRatingStageCreateView(LoginRequiredMixin, CreateView):
    model = FSI_CreditRating_Stage
    form_class = CreditRatingStageForm
    template_name = 'staging/creditrating_stage_form.html'
    success_url = reverse_lazy('creditrating_stage_list')

    def form_valid(self, form):
        try:
            # Set the created_by field to the currently logged-in user
            instance = form.save(commit=False)
            instance.created_by = self.request.user
            instance.save()

            # Log the creation in the AuditTrail
            AuditTrail.objects.create(
                user=self.request.user,
                model_name='FSI_CreditRating_Stage',
                action='create',
                object_id=instance.pk,
                change_description=f"Created Credit Rating Stage: Credit Rating - {instance.credit_rating}, Stage - {instance.stage}",
                timestamp=now(),
            )
            
            messages.success(self.request, "Credit rating successfully added!")
            return super().form_valid(form)
        except IntegrityError as e:
            # Handle unique constraint violations or other database integrity issues
            messages.error(self.request, f"Integrity error: {e}")
            return self.form_invalid(form)
        except Exception as e:
            # Handle any other unexpected errors
            messages.error(self.request, f"An unexpected error occurred: {e}")
            return self.form_invalid(form)

    def form_invalid(self, form):
        # Add a general error message for validation errors
        messages.error(self.request, "There were errors in the form. Please correct them below.")
        return super().form_invalid(form)

# Update View for editing a credit rating
class CreditRatingStageUpdateView(LoginRequiredMixin, UpdateView):
    model = FSI_CreditRating_Stage
    form_class = CreditRatingStageForm
    template_name = 'staging/creditrating_stage_form.html'
    success_url = reverse_lazy('creditrating_stage_list')

    def form_valid(self, form):
        try:
            # Set the updated_by field to the currently logged-in user
            instance = form.save(commit=False)
            instance.created_by = self.request.user
            instance.save()
            # Log the update in the AuditTrail
            AuditTrail.objects.create(
                user=self.request.user,
                model_name='FSI_CreditRating_Stage',
                action='update',
                object_id=instance.pk,
                change_description=f"Updated Credit Rating Stage: Credit Rating - {instance.credit_rating}, Stage - {instance.stage}",
                timestamp=now(),
            )

            messages.success(self.request, "Credit rating successfully updated!")
            return super().form_valid(form)
        except IntegrityError as e:
            # Handle unique constraint violations or other database integrity issues
            messages.error(self.request, f"Integrity error: {e}")
            return self.form_invalid(form)
        except Exception as e:
            # Handle any other unexpected errors
            messages.error(self.request, f"An unexpected error occurred: {e}")
            return self.form_invalid(form)

    def form_invalid(self, form):
        # Add a general error message for validation errors
        messages.error(self.request, "There were errors in the form. Please correct them below.")
        return super().form_invalid(form)

# Delete View for deleting a credit rating
# Delete View for deleting a credit rating
class CreditRatingStageDeleteView(LoginRequiredMixin, DeleteView):
    model = FSI_CreditRating_Stage
    template_name = 'staging/creditrating_stage_confirm_delete.html'
    success_url = reverse_lazy('creditrating_stage_list')

    def delete(self, request, *args, **kwargs):
        instance = self.get_object()
        try:
            # Log the deletion in the AuditTrail
            AuditTrail.objects.create(
                user=self.request.user,
                model_name='FSI_CreditRating_Stage',
                action='delete',
                object_id=instance.pk,
                change_description=f"Deleted Credit Rating Stage: Credit Rating - {instance.credit_rating}, Stage - {instance.stage}",
                timestamp=now(),
            )

            messages.success(self.request, "Credit rating successfully deleted!")
        except Exception as e:
            messages.error(self.request, f"Error logging deletion: {e}")

        return super().delete(request, *args, **kwargs)

##DPD staging

class DPDStageMappingListView(LoginRequiredMixin,ListView):
    model = FSI_DPD_Stage_Mapping
    template_name = 'staging/dpd_stage_mapping_list.html'
    context_object_name = 'dpd_mappings'
    paginate_by = 10

class DPDStageMappingCreateView(LoginRequiredMixin, CreateView):
    model = FSI_DPD_Stage_Mapping
    fields = ['payment_frequency', 'stage_1_threshold', 'stage_2_threshold', 'stage_3_threshold']
    template_name = 'staging/dpd_stage_mapping_form.html'
    success_url = reverse_lazy('dpd_stage_mapping_list')

    def form_valid(self, form):
        try:
            # Set the updated_by field to the currently logged-in user
            instance = form.save(commit=False)
            instance.created_by = self.request.user
            instance.save()
            # Log the creation in the AuditTrail
            AuditTrail.objects.create(
                user=self.request.user,
                model_name='FSI_DPD_Stage_Mapping',
                action='create',
                object_id=instance.pk,
                change_description=(
                    f"Created DPD Stage Mapping: Payment Frequency - {instance.payment_frequency}, "
                    f"Stage 1 Threshold - {instance.stage_1_threshold}, "
                    f"Stage 2 Threshold - {instance.stage_2_threshold}, "
                    f"Stage 3 Threshold - {instance.stage_3_threshold}"
                ),
                timestamp=now(),
            )

            messages.success(self.request, "DPD Stage Mapping successfully added!")
            return super().form_valid(form)
        except IntegrityError as e:
            # Handle unique constraint violations or other database integrity issues
            messages.error(self.request, f"Integrity error: {e}")
            return self.form_invalid(form)
        except Exception as e:
            # Handle any other unexpected errors
            messages.error(self.request, f"An unexpected error occurred: {e}")
            return self.form_invalid(form)

    def form_invalid(self, form):
        # Add a general error message for validation errors
        messages.error(self.request, "There were errors in the form. Please correct them below.")
        return super().form_invalid(form)

class DPDStageMappingUpdateView(LoginRequiredMixin, UpdateView):
    model = FSI_DPD_Stage_Mapping
    fields = ['payment_frequency', 'stage_1_threshold', 'stage_2_threshold', 'stage_3_threshold']
    template_name = 'staging/dpd_stage_mapping_form.html'
    success_url = reverse_lazy('dpd_stage_mapping_list')

    def form_valid(self, form):
        try:
            # Set the updated_by field to the currently logged-in user
            instance = form.save(commit=False)
            instance.created_by = self.request.user
            instance.save()
            # Log the update in the AuditTrail
            AuditTrail.objects.create(
                user=self.request.user,
                model_name='FSI_DPD_Stage_Mapping',
                action='update',
                object_id=instance.pk,
                change_description=(
                    f"Updated DPD Stage Mapping: Payment Frequency - {instance.payment_frequency}, "
                    f"Stage 1 Threshold - {instance.stage_1_threshold}, "
                    f"Stage 2 Threshold - {instance.stage_2_threshold}, "
                    f"Stage 3 Threshold - {instance.stage_3_threshold}"
                ),
                timestamp=now(),
            )
            messages.success(self.request, "DPD Stage Mapping successfully updated!")
            return super().form_valid(form)
        
        except IntegrityError as e:
            # Handle unique constraint violations or other database integrity issues
            messages.error(self.request, f"Integrity error: {e}")
            return self.form_invalid(form)
        except Exception as e:
            # Handle any other unexpected errors
            messages.error(self.request, f"An unexpected error occurred: {e}")
            return self.form_invalid(form)

    def form_invalid(self, form):
        # Add a general error message for validation errors
        messages.error(self.request, "There were errors in the form. Please correct them below.")
        return super().form_invalid(form)

class DPDStageMappingDeleteView(LoginRequiredMixin, DeleteView):
    model = FSI_DPD_Stage_Mapping
    template_name = 'staging/dpd_stage_mapping_confirm_delete.html'
    success_url = reverse_lazy('dpd_stage_mapping_list')

    def delete(self, request, *args, **kwargs):
        instance = self.get_object()
        try:
            # Log the deletion in the AuditTrail
            AuditTrail.objects.create(
                user=self.request.user,
                model_name='FSI_DPD_Stage_Mapping',
                action='delete',
                object_id=instance.pk,
                change_description=(
                    f"Deleted DPD Stage Mapping: Payment Frequency - {instance.payment_frequency}, "
                    f"Stage 1 Threshold - {instance.stage_1_threshold}, "
                    f"Stage 2 Threshold - {instance.stage_2_threshold}, "
                    f"Stage 3 Threshold - {instance.stage_3_threshold}"
                ),
                timestamp=now(),
            )

            messages.success(self.request, "DPD Stage Mapping successfully deleted!")
        except Exception as e:
            messages.error(self.request, f"Error logging deletion: {e}")

        return super().delete(request, *args, **kwargs)

#CoolingPeriodDefinition

class CoolingPeriodDefinitionListView(LoginRequiredMixin,ListView):
    model = CoolingPeriodDefinition
    template_name = 'staging/cooling_period_definition_list.html'
    context_object_name = 'cooling_periods'
    paginate_by = 10

class CoolingPeriodDefinitionCreateView(LoginRequiredMixin, CreateView):
    model = CoolingPeriodDefinition
    form_class = CoolingPeriodDefinitionForm
    template_name = 'staging/cooling_period_definition_form.html'
    success_url = reverse_lazy('cooling_period_definitions')

    def form_valid(self, form):
        try:
            # Set the updated_by field to the currently logged-in user
            instance = form.save(commit=False)
            instance.created_by = self.request.user
            instance.save()

            # Log the creation in the AuditTrail
            AuditTrail.objects.create(
                user=self.request.user,
                model_name='CoolingPeriodDefinition',
                action='create',
                object_id=instance.pk,
                change_description=(
                    f"Created Cooling Period Definition: Term Unit - {instance.v_amrt_term_unit}, "
                    f"Cooling Period Days - {instance.n_cooling_period_days}"
                ),
                timestamp=now(),
            )

            messages.success(self.request, "Cooling period definition successfully added!")
            return super().form_valid(form)
        
        except IntegrityError as e:
            # Handle unique constraint violations or other database integrity issues
            messages.error(self.request, f"Integrity error: {e}")
            return self.form_invalid(form)
        except Exception as e:
            # Handle any other unexpected errors
            messages.error(self.request, f"An unexpected error occurred: {e}")
            return self.form_invalid(form)

    def form_invalid(self, form):
        # Add a general error message for validation errors
        messages.error(self.request, "There were errors in the form. Please correct them below.")
        return super().form_invalid(form)

class CoolingPeriodDefinitionUpdateView(LoginRequiredMixin, UpdateView):
    model = CoolingPeriodDefinition
    form_class = CoolingPeriodDefinitionForm
    template_name = 'staging/cooling_period_definition_form.html'
    success_url = reverse_lazy('cooling_period_definitions')

    def form_valid(self, form):
        try:
            # Set the updated_by field to the currently logged-in user
            instance = form.save(commit=False)
            instance.created_by = self.request.user
            instance.save()
            # Log the update in the AuditTrail
            AuditTrail.objects.create(
                user=self.request.user,
                model_name='CoolingPeriodDefinition',
                action='update',
                object_id=instance.pk,
                change_description=(
                    f"Updated Cooling Period Definition: Term Unit - {instance.v_amrt_term_unit}, "
                    f"Cooling Period Days - {instance.n_cooling_period_days}"
                ),
                timestamp=now(),
            )

            messages.success(self.request, "Cooling period definition successfully updated!")
            return super().form_valid(form)
        
        except IntegrityError as e:
            # Handle unique constraint violations or other database integrity issues
            messages.error(self.request, f"Integrity error: {e}")
            return self.form_invalid(form)
        except Exception as e:
            # Handle any other unexpected errors
            messages.error(self.request, f"An unexpected error occurred: {e}")
            return self.form_invalid(form)

    def form_invalid(self, form):
        # Add a general error message for validation errors
        messages.error(self.request, "There were errors in the form. Please correct them below.")
        return super().form_invalid(form)

class CoolingPeriodDefinitionDeleteView(LoginRequiredMixin, DeleteView):
    model = CoolingPeriodDefinition
    template_name = 'staging/cooling_period_definition_confirm_delete.html'
    success_url = reverse_lazy('cooling_period_definitions')

    def delete(self, request, *args, **kwargs):
        instance = self.get_object()
        try:
            # Log the deletion in the AuditTrail
            AuditTrail.objects.create(
                user=self.request.user,
                model_name='CoolingPeriodDefinition',
                action='delete',
                object_id=instance.pk,
                change_description=(
                    f"Deleted Cooling Period Definition: Term Unit - {instance.v_amrt_term_unit}, "
                    f"Cooling Period Days - {instance.n_cooling_period_days}"
                ),
                timestamp=now(),
            )

            messages.success(self.request, "Cooling period definition successfully deleted!")
        except Exception as e:
            messages.error(self.request, f"Error logging deletion: {e}")

        return super().delete(request, *args, **kwargs)
    
class DimDelinquencyBandListView(LoginRequiredMixin,ListView):
    model = Dim_Delinquency_Band
    template_name = 'staging/dim_delinquency_band_list.html'
    context_object_name = 'delinquency_bands'
    paginate_by = 10

class DimDelinquencyBandCreateView(LoginRequiredMixin, CreateView):
    model = Dim_Delinquency_Band
    form_class = DimDelinquencyBandForm
    template_name = 'staging/dim_delinquency_band_form.html'
    success_url = reverse_lazy('dim_delinquency_band_list')

    def form_valid(self, form):
        try:

            # Set the updated_by field to the currently logged-in user
            instance = form.save(commit=False)
            instance.created_by = self.request.user
            instance.save()

            # Log the creation in the AuditTrail
            AuditTrail.objects.create(
                user=self.request.user,
                model_name='Dim_Delinquency_Band',
                action='create',
                object_id=instance.n_delq_band_code,
                change_description=(
                    f"Created Delinquency Band: Code - {instance.n_delq_band_code}, "
                    f"Description - {instance.v_delq_band_desc}, "
                    f"Lower Value - {instance.n_delq_lower_value}, "
                    f"Upper Value - {instance.n_delq_upper_value}"
                ),
                timestamp=now(),
            )

            messages.success(self.request, "Delinquency band successfully added!")
            return super().form_valid(form)
        
        except IntegrityError as e:
            # Handle unique constraint violations or other database integrity issues
            messages.error(self.request, f"Integrity error: {e}")
            return self.form_invalid(form)
        except Exception as e:
            # Handle any other unexpected errors
            messages.error(self.request, f"An unexpected error occurred: {e}")
            return self.form_invalid(form)

    def form_invalid(self, form):
        # Add a general error message for validation errors
        messages.error(self.request, "There were errors in the form. Please correct them below.")
        return super().form_invalid(form)

class DimDelinquencyBandUpdateView(LoginRequiredMixin, UpdateView):
    model = Dim_Delinquency_Band
    form_class = DimDelinquencyBandForm
    template_name = 'staging/dim_delinquency_band_form.html'
    success_url = reverse_lazy('dim_delinquency_band_list')

    def form_valid(self, form):
        try:

            # Set the updated_by field to the currently logged-in user
            instance = form.save(commit=False)
            instance.created_by = self.request.user
            instance.save()

            # Log the update in the AuditTrail
            AuditTrail.objects.create(
                user=self.request.user,
                model_name='Dim_Delinquency_Band',
                action='update',
                object_id=instance.n_delq_band_code,
                change_description=(
                    f"Updated Delinquency Band: Code - {instance.n_delq_band_code}, "
                    f"Description - {instance.v_delq_band_desc}, "
                    f"Lower Value - {instance.n_delq_lower_value}, "
                    f"Upper Value - {instance.n_delq_upper_value}"
                ),
                timestamp=now(),
            )

            messages.success(self.request, "Delinquency band successfully updated!")
            return super().form_valid(form)
        
        except IntegrityError as e:
            # Handle unique constraint violations or other database integrity issues
            messages.error(self.request, f"Integrity error: {e}")
            return self.form_invalid(form)
        except Exception as e:
            # Handle any other unexpected errors
            messages.error(self.request, f"An unexpected error occurred: {e}")
            return self.form_invalid(form)

    def form_invalid(self, form):
        # Add a general error message for validation errors
        messages.error(self.request, "There were errors in the form. Please correct them below.")
        return super().form_invalid(form)

class DimDelinquencyBandDeleteView(LoginRequiredMixin, DeleteView):
    model = Dim_Delinquency_Band
    template_name = 'staging/dim_delinquency_band_confirm_delete.html'
    success_url = reverse_lazy('dim_delinquency_band_list')

    def delete(self, request, *args, **kwargs):
        instance = self.get_object()
        try:
            # Log the deletion in the AuditTrail
            AuditTrail.objects.create(
                user=self.request.user,
                model_name='Dim_Delinquency_Band',
                action='delete',
                object_id=instance.n_delq_band_code,
                change_description=(
                    f"Deleted Delinquency Band: Code - {instance.n_delq_band_code}, "
                    f"Description - {instance.v_delq_band_desc}, "
                    f"Lower Value - {instance.n_delq_lower_value}, "
                    f"Upper Value - {instance.n_delq_upper_value}"
                ),
                timestamp=now(),
            )

            messages.success(self.request, "Delinquency band successfully deleted!")
        except Exception as e:
            messages.error(self.request, f"Error logging deletion: {e}")

        return super().delete(request, *args, **kwargs)
class CreditRatingCodeBandListView(LoginRequiredMixin,ListView):
    model = Credit_Rating_Code_Band
    template_name = 'staging/credit_rating_code_band_list.html'
    context_object_name = 'rating_codes'
    paginate_by = 10

class CreditRatingCodeBandCreateView(LoginRequiredMixin, CreateView):
    model = Credit_Rating_Code_Band
    form_class = CreditRatingCodeBandForm
    template_name = 'staging/credit_rating_code_band_form.html'
    success_url = reverse_lazy('credit_rating_code_band_list')

    def form_valid(self, form):
        try:
            # Set the updated_by field to the currently logged-in user
            instance = form.save(commit=False)
            instance.created_by = self.request.user
            instance.save()

            # Log the creation in the AuditTrail
            AuditTrail.objects.create(
                user=self.request.user,
                model_name='Credit_Rating_Code_Band',
                action='create',
                object_id=instance.v_rating_code,
                change_description=(
                    f"Created Credit Rating Code Band: Code - {instance.v_rating_code}, "
                    f"Description - {instance.v_rating_desc}"
                ),
                timestamp=now(),
            )

            messages.success(self.request, "Credit rating code successfully added!")
            return super().form_valid(form)
        
        except IntegrityError as e:
            # Handle unique constraint violations or other database integrity issues
            messages.error(self.request, f"Integrity error: {e}")
            return self.form_invalid(form)
        except Exception as e:
            # Handle any other unexpected errors
            messages.error(self.request, f"An unexpected error occurred: {e}")
            return self.form_invalid(form)

    def form_invalid(self, form):
        # Add a general error message for validation errors
        messages.error(self.request, "There were errors in the form. Please correct them below.")
        return super().form_invalid(form)

class CreditRatingCodeBandUpdateView(LoginRequiredMixin, UpdateView):
    model = Credit_Rating_Code_Band
    form_class = CreditRatingCodeBandForm
    template_name = 'staging/credit_rating_code_band_form.html'
    success_url = reverse_lazy('credit_rating_code_band_list')

    def form_valid(self, form):
        try:
            # Set the updated_by field to the currently logged-in user
            instance = form.save(commit=False)
            instance.created_by = self.request.user
            instance.save()
            # Log the update in the AuditTrail
            AuditTrail.objects.create(
                user=self.request.user,
                model_name='Credit_Rating_Code_Band',
                action='update',
                object_id=instance.v_rating_code,
                change_description=(
                    f"Updated Credit Rating Code Band: Code - {instance.v_rating_code}, "
                    f"Description - {instance.v_rating_desc}"
                ),
                timestamp=now(),
            )

            messages.success(self.request, "redit rating code successfully updated!")
            return super().form_valid(form)
        
        except IntegrityError as e:
            # Handle database integrity errors, like duplicate entries
            messages.error(self.request, f"Integrity error: {e}")
            return self.form_invalid(form)
        except Exception as e:
            # Handle any other exceptions
            messages.error(self.request, f"An unexpected error occurred: {e}")
            return self.form_invalid(form)

class CreditRatingCodeBandDeleteView(LoginRequiredMixin, DeleteView):
    model = Credit_Rating_Code_Band
    template_name = 'staging/credit_rating_code_band_confirm_delete.html'
    success_url = reverse_lazy('credit_rating_code_band_list')

    def delete(self, request, *args, **kwargs):
        instance = self.get_object()
        try:
            # Log the deletion in the AuditTrail
            AuditTrail.objects.create(
                user=self.request.user,
                model_name='Credit_Rating_Code_Band',
                action='delete',
                object_id=instance.v_rating_code,
                change_description=(
                    f"Deleted Credit Rating Code Band: Code - {instance.v_rating_code}, "
                    f"Description - {instance.v_rating_desc}"
                ),
                timestamp=now(),
            )

            messages.success(self.request, "Credit rating code successfully deleted!")
        except Exception as e:
            messages.error(self.request, f"Error logging deletion: {e}")

        return super().delete(request, *args, **kwargs)
    


#stage reassignment
# Map for stage descriptions based on the selected stage key
# Map for stage descriptions based on the selected stage key
STAGE_DESCRIPTION_MAP = {
    1: 'Stage 1',
    2: 'Stage 2',
    3: 'Stage 3',
}



@login_required
def stage_reassignment(request): 
    filter_form = StageReassignmentFilterForm(request.GET or None)
    records = None

    # Fetch available Reporting Dates in descending order
    fic_mis_dates = FCT_Reporting_Lines.objects.values_list('fic_mis_date', flat=True).distinct().order_by('-fic_mis_date')

    try:
        if filter_form.is_valid():
            fic_mis_date = filter_form.cleaned_data.get('fic_mis_date')
            n_account_number = filter_form.cleaned_data.get('n_account_number')
            n_cust_ref_code = filter_form.cleaned_data.get('n_cust_ref_code')

            # Filter logic
            records = FCT_Reporting_Lines.objects.filter(fic_mis_date=fic_mis_date, n_account_number=n_account_number)
            if n_cust_ref_code:
                records = records.filter(n_cust_ref_code=n_cust_ref_code)

            # Check if records are empty and set an appropriate message
            if not records.exists():
                messages.warning(request, "No data records were retrieved. The query condition returned zero records.")

        if request.method == "POST":
            stage_form = StageReassignmentForm(request.POST)
            if stage_form.is_valid():
                with transaction.atomic():  # Use atomic transaction to ensure database integrity
                    # Loop through the records and update the stages
                    for record in records:
                        stage_key = request.POST.get(f"n_curr_ifrs_stage_skey_{record.n_account_number}")
                        if stage_key:
                            previous_stage = record.n_stage_descr
                            record.n_curr_ifrs_stage_skey = int(stage_key)
                            record.n_stage_descr = STAGE_DESCRIPTION_MAP.get(int(stage_key))
                            record.save()
                            # Log the stage reassignment in the AuditTrail
                            AuditTrail.objects.create(
                                user=request.user,
                                model_name='FCT_Reporting_Lines',
                                action='update',
                                object_id=record.n_account_number,
                                change_description=(
                                    f"Reassigned stage for account {record.n_account_number}: "
                                    f"From Stage - {previous_stage}, To Stage - {record.n_stage_descr}"
                                ),
                                timestamp=now(),
                            )

                    messages.success(request, "Stages reassigned successfully!")
            else:
                messages.error(request, "Invalid data provided.")

    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")

    # Use this check instead of is_ajax()
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        # Return the partial HTML if the request is via AJAX
        return render(request, 'staging/partials/stage_reassignment_table.html', {
            'records': records,
            'filter_form': filter_form,
            'stage_form': StageReassignmentForm()
        })

    # For non-AJAX requests, return the full page
    return render(request, 'staging/stage_reassignment.html', {
        'records': records,
        'filter_form': filter_form,
        'stage_form': StageReassignmentForm(),
        'fic_mis_dates': fic_mis_dates
    })


