from django.shortcuts import render, get_object_or_404, redirect
from django.forms import inlineformset_factory
from django.db import transaction
import threading
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from IFRS9.models import Process, RunProcess,Function,FunctionExecutionStatus
from Users.models import AuditTrail  # Import AuditTrail model
from django.utils.timezone import now  # For timestamping
from scylla_ifrs9_postgre.forms import ProcessForm, RunProcessForm
from django.db import transaction
from django.db.models import Q
from django.db.models import Count
from django.http import JsonResponse
from django.template.loader import render_to_string
from django.db.models import Max
from django.db.models import Min
from datetime import datetime
from Users.models import AuditTrail  # Import AuditTrail model
from django.utils.timezone import now  # For timestamping
from django.utils import timezone
from django.utils.module_loading import import_string  # Used for dynamic function calling
from django.core.paginator import Paginator
from openpyxl.utils import get_column_letter
import csv
import openpyxl
from django.http import JsonResponse, HttpResponse
from ..Functions.cashflow import *
from ..Functions.calculate_cash_flows_ead import *
from ..Functions.calculate_cash_flows_exposure import *
from ..Functions.cal_provision_matrix import *
from ..Functions.pd_interpolation import *
from ..Functions.run_vasicek_model import *
from ..Functions.populate_stg_determination import *
from ..Functions.determine_stage import *
from ..Functions.cooling_period import *
from ..Functions.update_stage_determination import *
from ..Functions.calculate_lgd import *
from ..Functions.cal_hist_lgd_loss_rate_dpd import *
from ..Functions.run_frye_jacobs_lgd import *
from ..Functions.assign_acc_pd_level import *
from ..Functions.assign_acc_pd_term_level import *
from ..Functions.populate_cashflows import *
from ..Functions.pd_cumulative_term_str import *
from ..Functions.calculate_fct_accrued_interest_and_ead import *
from ..Functions.calculate_fct_EAD_using_accrued import *
from ..Functions.calculate_fct_EAD_using_cashflows_pv import *
from ..Functions.calculate_eir import *
from ..Functions.calculate_eir_using_cashflows import *
from ..Functions.update_fin_cashflw import *
from ..Functions.calculate_cash_flow_rate_and_amount1 import *
from ..Functions.cal_cf_lgd_using_collateral import *
from ..Functions.cal_periodic_discount_Rate2 import *
from ..Functions.cal_exp_cash_n_cash_shortfall3 import *
from ..Functions.cal_forward_exposure4 import *
from ..Functions.calculate_marginal_pd import *
from ..Functions.populate_reporting_table import *
from ..Functions.calculate_ecl import *
from ..Functions.cal_reporting_currency import *
from django.db.models import Min, Max


@login_required

def operations_view(request):
    return render(request, 'operations/operations.html')



# List all processes
@login_required
@permission_required('IFRS9.view_process', raise_exception=True)
def process_list(request):
    processes = Process.objects.all()
    return render(request, 'operations/process_list.html', {'processes': processes})

# View the details of a specific process, including its associated functions and execution order
@login_required
@permission_required('IFRS9.view_process', raise_exception=True)
def process_detail(request, process_id):
    process = get_object_or_404(Process, id=process_id)
    run_processes = RunProcess.objects.filter(process=process).order_by('order')  # Fetch functions in order
    return render(request, 'operations/process_detail.html', {'process': process, 'run_processes': run_processes})


# Create or edit a process
@login_required
@permission_required('IFRS9.add_process', raise_exception=True)
def create_process(request, process_id=None):
    """
    View to add or edit a process and its corresponding run processes.
    If process_id is provided, it's an edit; otherwise, it's an add.
    """
    RunProcessFormSet = inlineformset_factory(Process, RunProcess, form=RunProcessForm, extra=1, can_delete=True)

    
    process = Process()
    form_title = 'Create Process'

    if request.method == 'POST':
        form = ProcessForm(request.POST, instance=process)
        formset = RunProcessFormSet(request.POST)

        # Print the formset POST data for debugging
        print("Formset POST Data:", request.POST)

        if form.is_valid():
            try:
                with transaction.atomic():
                    process = form.save(commit=False)
                    process.created_by = request.user
                    process.save()  # Save the parent process object
                
                    # Log the creation in the AuditTrail
                    AuditTrail.objects.create(
                        user=request.user,
                        model_name='Process',
                        action='create',
                        object_id=process.id,
                        change_description=f"Created process: {process.process_name}",
                        timestamp=now(),
                    )

                    # Handle RunProcess formset
                    for form in formset.forms:
                        
                        run_process = form.save(commit=False)
                        run_process.process = process
                        
                        # Log the creation of RunProcess in the AuditTrail
                        AuditTrail.objects.create(
                            user=request.user,
                            model_name='RunProcess',
                            action='create',
                            object_id=run_process.id,
                            change_description=(
                                f"Added function {run_process.function.function_name} "
                                f"to process {process.process_name} with order {run_process.order}"
                            ),
                            timestamp=now(),
                        )


                    # Handle multiple function and order values
                    for form in formset.forms:
                        functions = request.POST.getlist(form.add_prefix('function'))  # Fetch as list
                        orders = request.POST.getlist(form.add_prefix('order'))  # Fetch as list

                        print(f"Processing form: functions={functions}, orders={orders}")

                        # Loop over the multiple values and save each pair separately
                        if len(functions) == len(orders):
                            for function_id, order in zip(functions, orders):
                                if function_id and order:
                                    # Convert the function_id to a Function instance
                                    function_instance = Function.objects.get(pk=function_id)
                                    
                                    RunProcess.objects.create(
                                        process=process,
                                        function=function_instance,  # Save the Function instance
                                        order=order
                                    )
                                    
                                    print(f"Saved: function={function_instance}, order={order}")
                        else:
                            # Handle single function and order values
                            if form.cleaned_data and not form.cleaned_data.get('DELETE', False):
                                function_instance = form.cleaned_data.get('function')
                                RunProcess.objects.create(
                                    process=process,
                                    function=function_instance,  # Save the Function instance
                                    order=form.cleaned_data.get('order')
                                )
                                print(f"Saved single: function={form.cleaned_data.get('function')}, order={form.cleaned_data.get('order')}")

                    messages.success(request, f'Process {"created" if not process_id else "updated"} successfully.')
                    return redirect('process_list')
            except Exception as e:
                messages.error(request, f"Error saving process: {str(e)}")
                print(f"Error: {e}")
        else:
            print("Form Errors:", form.errors)
            print("Formset Errors:", formset.errors)
            messages.error(request, "Please correct the errors below.")
    else:
        form = ProcessForm(instance=process)
        formset = RunProcessFormSet(instance=process)

    return render(request, 'operations/create_process.html', {
        'form': form,
        'formset': formset,
        'title': form_title,
    })

# Edit an existing process
@login_required
@permission_required('IFRS9.change_process', raise_exception=True)
def edit_process(request, process_id):
    """
    View to edit an existing process and its corresponding run processes.
    """
    # Define the formset for RunProcess with can_delete=True
    RunProcessFormSet = inlineformset_factory(
        Process,
        RunProcess,
        form=RunProcessForm,
        extra=1,
        can_delete=True,  # Enables deletion functionality
    )

    # Fetch the process object to edit or return 404 if not found
    process = get_object_or_404(Process, id=process_id)
    form_title = "Edit Process"

    if request.method == "POST":
        form = ProcessForm(request.POST, instance=process)
        formset = RunProcessFormSet(request.POST, instance=process)

        # Debugging: Log errors for clarity
        print("Edit Form POST Data:", request.POST)
        print("Edit Form Errors:", form.errors)
        print("Edit Formset Errors:", formset.errors)

        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    # Save the process
                    previous_name = process.process_name
                    process = form.save(commit=False)
                    process.created_by = request.user
                    process.save()

                    # Log the update in the AuditTrail
                    AuditTrail.objects.create(
                        user=request.user,
                        model_name='Process',
                        action='update',
                        object_id=process.id,
                        change_description=(
                            f"Updated process: from {previous_name} to {process.process_name}"
                        ),
                        timestamp=now(),
                    )
                    # Save the RunProcess instances
                    for form in formset.forms:
                        if form.cleaned_data and not form.cleaned_data.get('DELETE', False):
                            run_process = form.save(commit=False)
                            run_process.process = process
                            

                            # Log updates or additions of RunProcess
                            AuditTrail.objects.create(
                                user=request.user,
                                model_name='RunProcess',
                                action='update' if run_process.id else 'create',
                                object_id=run_process.id,
                                change_description=(
                                    f"{'Updated' if run_process.id else 'Added'} function {run_process.function.function_name} "
                                    f"to process {process.process_name} with order {run_process.order}"
                                ),
                                timestamp=now(),
                            )

                    # Save the formset (RunProcess instances), including deletions
                    formset.save()

                    messages.success(request, "Process updated successfully.")
                    return redirect("process_list")
            except Exception as e:
                messages.error(request, f"Error updating process: {str(e)}")
                print(f"Error: {e}")
        else:
            messages.error(
                request,
                "There were errors in the form. Please correct them and try again.",
            )
    else:
        form = ProcessForm(instance=process)
        formset = RunProcessFormSet(instance=process)

    return render(
        request,
        "operations/edit_process.html",
        {
            "form": form,
            "formset": formset,
            "title": form_title,
            "process": process,
        },
    )


@login_required
@permission_required('IFRS9.delete_process', raise_exception=True)
def delete_process(request, process_id):
    process = get_object_or_404(Process, id=process_id)
    if request.method == 'POST':
        try:
            # Log the deletion in the AuditTrail
            AuditTrail.objects.create(
                user=request.user,
                model_name='Process',
                action='delete',
                object_id=process.id,
                change_description=f"Deleted process: {process.process_name}",
                timestamp=now(),
            )

            process.delete()
            messages.success(request, 'Process deleted successfully.')
        except Exception as e:
            messages.error(request, f'Error deleting process: {e}')

        return redirect('process_list')  # Go back to the list after deletion

    # If GET, just show the "Are you sure?" page
    return render(request, 'operations/delete_process.html', {'process': process})
##############################################################################################3
# Display and search for processes
@login_required
def execute_process_view(request):
    query = request.GET.get('search', '')
    processes = Process.objects.filter(Q(process_name__icontains=query))

    return render(request, 'operations/execute_process.html', {
        'processes': processes,
        'query': query,
    })

# Handle execution
# Function to generate the process run ID and count

running_threads = {}
cancel_flags = {}


def generate_process_run_id(process, execution_date):
    """
    Generate a process_run_id in the format 'process_id_execution_date_run_number'.
    """
    # Format the execution date as YYYYMMDD
    execution_date_str = execution_date.strftime('%Y%m%d')
    
    # Base run ID: process_id + execution_date
    base_run_id = f"{process.process_name}_{execution_date_str}"
    
    # Check the database for existing entries with the same base_run_id
    existing_runs = FunctionExecutionStatus.objects.filter(process_run_id__startswith=base_run_id).order_by('-run_count')
    
    # Determine the next run count
    if existing_runs.exists():
        last_run_count = existing_runs[0].run_count
        next_run_count = last_run_count + 1
    else:
        next_run_count = 1
    
    # Generate the full process_run_id with the next run count
    process_run_id = f"{base_run_id}_{next_run_count}"
    
    return process_run_id, next_run_count


# Background function for running the process
# Background function for running the process
def execute_functions_in_background(function_status_entries, process_run_id, mis_date):
    for status_entry in function_status_entries:

        if cancel_flags.get(process_run_id):  # Check if cancellation was requested
            status_entry.status = 'Cancelled'
            status_entry.execution_end_date = timezone.now()
            status_entry.duration = status_entry.execution_end_date - status_entry.execution_start_date
            status_entry.save()
            print(f"Process {process_run_id} was cancelled.")
            break  # Stop execution if cancelled

        function_name = status_entry.function.function_name
        print(f"Preparing to execute function: {function_name}")

        # Set the function status to "Ongoing" and record the start date
        status_entry.status = 'Ongoing'
        status_entry.execution_start_date = timezone.now()  # Start time for the function
        status_entry.save()
        print(f"Function {function_name} marked as Ongoing.")

        # Execute the function
        try:
            if function_name in globals():
                print(f"Executing function: {function_name} with date {mis_date}")
                result = globals()[function_name](mis_date)  # Execute the function and capture the return value
                
                # Update status and end date based on the return value
                status_entry.execution_end_date = timezone.now()  # End time for the function
                if result == 1 or result == '1':
                    status_entry.status = 'Success'
                    print(f"Function {function_name} executed successfully.")
                elif result == 0 or result == '0':
                    status_entry.status = 'Failed'
                    print(f"Function {function_name} execution failed.")
                    status_entry.save()
                    break  # Stop execution if the function fails
                else:
                    status_entry.status = 'Failed'
                    print(f"Unexpected return value {result} from function {function_name}.")
                    status_entry.save()
                    break  # Stop execution for any unexpected result
            else:
                status_entry.status = 'Failed'
                print(f"Function {function_name} not found in the global scope.")
                status_entry.execution_end_date = timezone.now()
                status_entry.save()
                break  # Stop execution if the function is not found

        except Exception as e:
            status_entry.status = 'Failed'
            status_entry.execution_end_date = timezone.now()
            print(f"Error executing {function_name}: {e}")
            status_entry.save()
            break  # Stop execution if any function throws an exception

        # Calculate duration
        if status_entry.execution_start_date and status_entry.execution_end_date:
            status_entry.duration = status_entry.execution_end_date - status_entry.execution_start_date

        # Save the final status and duration
        status_entry.save()
        print(f"Updated FunctionExecutionStatus for {function_name} to {status_entry.status}")


@login_required
@permission_required('IFRS9.can_execute_run',raise_exception=True)
def run_process_execution(request):
    if request.method == 'POST':
        process_id = request.POST.get('process_id')
        selected_function_ids = request.POST.getlist('selected_functions')
        
        # Parse the execution date
        mis_date = request.POST.get('execution_date')
        execution_date = datetime.strptime(mis_date, '%Y-%m-%d')
        print(f"Execution date received: {mis_date}")
        
        # Retrieve the selected process
        process = get_object_or_404(Process, id=process_id)
        print(f"Process selected: {process.process_name} (ID: {process.id})")
        
        # Fetch the RunProcess records in order of their execution (by 'order' field)
        # run_processes = RunProcess.objects.filter(id__in=selected_function_ids).order_by('order')
        run_processes = RunProcess.objects.filter(process=process).order_by('order')
        print(f"Number of selected functions to execute: {run_processes.count()}")

        # Generate the process_run_id and run_count
        process_run_id, run_count = generate_process_run_id(process, execution_date)
        print(f"Generated process_run_id: {process_run_id}, run_count: {run_count}")

        # Save all functions as "Pending"
        function_status_entries = []
        for run_process in run_processes:
            status_entry = FunctionExecutionStatus.objects.create(
                process=process,
                function=run_process.function,
                reporting_date=mis_date,  # Use the original string date for the execution status
                status='Pending',  # Initially marked as "Pending"
                process_run_id=process_run_id,
                run_count=run_count,
                execution_order=run_process.order,
                created_by=request.user

            )
            function_status_entries.append(status_entry)
            print(f"Function {run_process.function.function_name} marked as Pending.")

        # Redirect to the monitoring page so the user can see the function statuses
        response = redirect('monitor_specific_process', process_run_id=process_run_id)

        # Execute functions in the background (thread)
        execution_thread = threading.Thread(target=execute_functions_in_background, args=(function_status_entries, process_run_id, mis_date))
        execution_thread.start()

        return response  # Redirects immediately while the background task executes
        

@login_required
def get_process_functions(request, process_id):
    process = get_object_or_404(Process, id=process_id)
    functions_html = render_to_string('operations/_functions_list.html', {'run_processes': process.run_processes.all()})
    return JsonResponse({'html': functions_html})



# Monitor running process page
@login_required
def monitor_running_process_view(request):
    # Fetch distinct reporting dates and order by date descending
    available_dates = FunctionExecutionStatus.objects.order_by('-reporting_date').values_list('reporting_date', flat=True).distinct()

    # Get the selected date from the request
    selected_date = request.GET.get('selected_date', '')

    # Filter processes based on the selected date and ensure uniqueness by using annotation
    processes = []
    if selected_date:
        processes = (
            FunctionExecutionStatus.objects.filter(reporting_date=selected_date)
            .values('process__process_name', 'process_run_id','created_by__email')
            .annotate(
                latest_run=Max('process_run_id'),  # Latest run ID
                start_time=Min('execution_start_date'),  # Earliest start time
                end_time=Max('execution_end_date')  # Latest end time
            )
        )

        # Calculate overall status and duration for each process
        for process in processes:
            process_run_id = process['process_run_id']
            function_statuses = FunctionExecutionStatus.objects.filter(process_run_id=process_run_id)

            # Determine the overall status based on the function statuses
            if function_statuses.filter(status='Failed').exists():
                process['overall_status'] = 'Failed'
            elif function_statuses.filter(status='Cancelled').exists():
                process['overall_status'] = 'Cancelled'
            elif function_statuses.filter(status='Ongoing').exists():
                process['overall_status'] = 'Ongoing'
            else:
                process['overall_status'] = 'Success'
            

            # Calculate duration if start and end times are available
            start_time = process['start_time']
            end_time = process['end_time']
            if start_time and end_time:
                duration = end_time - start_time
                process['duration'] = duration.total_seconds() / 60  # Convert to minutes
            else:
                process['duration'] = None

    context = {
        'selected_date': selected_date,
        'processes': processes,
        'available_dates': available_dates,

    }
    return render(request, 'operations/monitor_running_process.html', context)


@login_required
def monitor_specific_process(request, process_run_id):
    # Fetch the specific process run by its ID
    process_statuses = FunctionExecutionStatus.objects.filter(process_run_id=process_run_id)

    context = {
        'process_statuses': process_statuses,
        'process_run_id': process_run_id,
    }
    return render(request, 'operations/monitor_specific_process.html', context)

@login_required
def get_updated_status_table(request):
    process_run_id = request.GET.get('process_run_id')
    
    # Get all statuses related to the process_run_id
    function_statuses = FunctionExecutionStatus.objects.filter(process_run_id=process_run_id).order_by('execution_order')
    
    # Add total_seconds to each status entry
    for status in function_statuses:
        if isinstance(status.duration, timedelta):
            # Calculate total seconds and assign it as an additional attribute for template access
            status.total_seconds = status.duration.total_seconds()
        else:
            status.total_seconds = None

    # Return the partial template with the updated table
    return render(request, 'operations/status_table.html', {'function_statuses': function_statuses})

@login_required
def get_process_function_status(request, process_run_id):
    # Get all statuses related to the process_run_id
    run_processes = FunctionExecutionStatus.objects.filter(process_run_id=process_run_id).order_by('execution_order')

    # Compute duration details for each process
    for process in run_processes:
        if isinstance(process.duration, timedelta):
            total_seconds = process.duration.total_seconds()
            process.total_seconds = total_seconds
            process.minutes = int(total_seconds // 60)
            process.remaining_seconds = int(total_seconds % 60)
        else:
            process.total_seconds = None
            process.minutes = None
            process.remaining_seconds = None

    # Render the updated table to HTML
    functions_html = render_to_string('operations/_function_status_list.html', {'run_processes': run_processes})
    return JsonResponse({'html': functions_html})



@login_required
def running_processes_view(request):
    # Fetch all ongoing (running) processes from the FunctionExecutionStatus model
    running_processes = FunctionExecutionStatus.objects.filter(status='Ongoing')

    context = {
        'running_processes': running_processes
    }

    return render(request, 'operations/running_processes.html', context)

# Updated function to handle cancellation request
@login_required
def cancel_running_process(request, process_run_id):
    # Check if the process is running
    try:
        # Get all functions related to the given process_run_id
        functions = FunctionExecutionStatus.objects.filter(
            process_run_id=process_run_id,
            status__in=['Pending', 'Ongoing']
        )
        
        if functions.exists():
            # Update the status of all "Pending" and "Ongoing" functions to "Cancelled"
            functions.update(status='Cancelled')
            messages.success(request, f"Process '{process_run_id}' and all pending functions have been cancelled.")
        else:
            messages.info(request, f"No running process found with the given ID '{process_run_id}'.")
    
    except FunctionExecutionStatus.DoesNotExist:
        messages.error(request, f"An error occurred while cancelling the process '{process_run_id}'.")

    return redirect('running_processes')  # Redirect to the running processes list


@login_required
def data_quality_check(request):
    """
    View for the Data Quality Check page.
    """
    return render(request, 'operations/data_quality_check.html')


@login_required
def check_missing_customers(request):
    """
    View to check for missing customers with AJAX support for pagination and Excel download.
    """
    # AJAX call to fetch fic_mis_dates
    if request.headers.get("x-requested-with") == "XMLHttpRequest" and "fic_mis_date" not in request.GET:
        fic_mis_dates = (
            Ldn_Financial_Instrument.objects.values_list("fic_mis_date", flat=True)
            .distinct()
            .order_by("-fic_mis_date")
        )
        return JsonResponse({"fic_mis_dates": list(fic_mis_dates)})

    # Excel download
    if request.GET.get("download") == "true":
        fic_mis_date = request.GET.get("fic_mis_date")
        if fic_mis_date:
            missing_customers = Ldn_Financial_Instrument.objects.filter(
                fic_mis_date=fic_mis_date
            ).exclude(
                v_cust_ref_code__in=Ldn_Customer_Info.objects.values_list("v_party_id", flat=True)
            ).values("v_cust_ref_code", "v_account_number", "fic_mis_date")

            # Create an Excel file
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Missing Customers"

            # Add headers
            headers = ["Customer Reference Code", "Account Number", "Reporting Date"]
            sheet.append(headers)

            # Add data rows
            for customer in missing_customers:
                sheet.append(
                    [
                        customer["v_cust_ref_code"],
                        customer["v_account_number"],
                        customer["fic_mis_date"],
                    ]
                )

            # Save the workbook to the HTTP response
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f'attachment; filename="missing_customers_{fic_mis_date}.xlsx"'
            workbook.save(response)  # Save directly to response
            return response
        
    # AJAX call to fetch paginated missing customers
    if request.headers.get("x-requested-with") == "XMLHttpRequest" and "fic_mis_date" in request.GET:
        fic_mis_date = request.GET.get("fic_mis_date")
        page = int(request.GET.get("page", 1))
        rows_per_page = int(request.GET.get("rows_per_page", 100))

        if fic_mis_date:
            missing_customers = Ldn_Financial_Instrument.objects.filter(
                fic_mis_date=fic_mis_date
            ).exclude(
                v_cust_ref_code__in=Ldn_Customer_Info.objects.values_list("v_party_id", flat=True)
            ).values("v_cust_ref_code", "v_account_number", "fic_mis_date")

            paginator = Paginator(missing_customers, rows_per_page)
            paginated_data = list(paginator.get_page(page).object_list)
            return JsonResponse(
                {
                    "data": paginated_data,
                    "total_pages": paginator.num_pages,
                    "current_page": page,
                    "total_records": paginator.count,
                }
            )
        else:
            return JsonResponse(
                {"data": [], "total_pages": 0, "current_page": 1, "total_records": 0}
            )

    

    # Render initial template
    return render(request, "operations/check_missing_customers.html", {})

@login_required
def check_missing_products(request):
    """
    View to check for missing products based on the given fic_mis_date.
    It compares v_prod_code in Ldn_Financial_Instrument with v_prod_code in Ldn_Bank_Product_Info.
    """

    # AJAX call to fetch fic_mis_dates
    if request.headers.get("x-requested-with") == "XMLHttpRequest" and "fic_mis_date" not in request.GET:
        fic_mis_dates = (
            Ldn_Financial_Instrument.objects.values_list("fic_mis_date", flat=True)
            .distinct()
            .order_by("-fic_mis_date")
        )
        return JsonResponse({"fic_mis_dates": list(fic_mis_dates)})

    # Excel download
    if request.GET.get("download") == "true":
        fic_mis_date = request.GET.get("fic_mis_date")
        if fic_mis_date:
            # Fetch distinct product codes
            missing_products = (
                Ldn_Financial_Instrument.objects.filter(
                    fic_mis_date=fic_mis_date
                )
                .exclude(
                    v_prod_code__in=Ldn_Bank_Product_Info.objects.values_list("v_prod_code", flat=True)
                )
                .values("v_prod_code", "fic_mis_date")
                .distinct()  # Ensure distinct product codes
                )

            # Create an Excel workbook
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Missing Products"

            # Add headers
            headers = ["Product Code", "Reporting Date"]
            sheet.append(headers)

            # Add data rows
            for product in missing_products:
                sheet.append(
                    [
                        product["v_prod_code"],
                        product["fic_mis_date"],
                    ]
                )

            # Save the workbook to the HTTP response
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f'attachment; filename="missing_products_{fic_mis_date}.xlsx"'
            workbook.save(response)  # Save directly to response
            return response

    # AJAX call to fetch paginated missing products
    if request.headers.get("x-requested-with") == "XMLHttpRequest" and "fic_mis_date" in request.GET:
        fic_mis_date = request.GET.get("fic_mis_date")
        page = int(request.GET.get("page", 1))
        rows_per_page = int(request.GET.get("rows_per_page", 100))

        if fic_mis_date:
            missing_products = (
                    Ldn_Financial_Instrument.objects.filter(
                        fic_mis_date=fic_mis_date
                    )
                    .exclude(
                        v_prod_code__in=Ldn_Bank_Product_Info.objects.values_list("v_prod_code", flat=True)
                    )
                    .values("v_prod_code", "fic_mis_date")
                    .distinct()  # Ensure distinct product codes
                                )


            paginator = Paginator(missing_products, rows_per_page)
            paginated_data = list(paginator.get_page(page).object_list)
            return JsonResponse(
                {
                    "data": paginated_data,
                    "total_pages": paginator.num_pages,
                    "current_page": page,
                    "total_records": paginator.count,
                }
            )
        else:
            return JsonResponse(
                {"data": [], "total_pages": 0, "current_page": 1, "total_records": 0}
            )

    # Render the initial template
    return render(request, "operations/check_missing_products.html", {})


@login_required
def check_cashflow_data(request):
    """
    View to check cashflow data quality based on provided rules.
    """
    if request.headers.get("x-requested-with") == "XMLHttpRequest" and "fic_mis_date" not in request.GET:
        fic_mis_dates = (
            Ldn_Financial_Instrument.objects.values_list("fic_mis_date", flat=True)
            .distinct()
            .order_by("-fic_mis_date")
        )
        return JsonResponse({"fic_mis_dates": list(fic_mis_dates)})

    fic_mis_date = request.GET.get("fic_mis_date")
    rows_per_page = int(request.GET.get("rows_per_page", 10))
    download = request.GET.get("download", False)

    if fic_mis_date:
        cashflow_data = Ldn_Financial_Instrument.objects.filter(
            fic_mis_date=fic_mis_date
        ).values(
            "v_account_number",
            "d_maturity_date",
            "d_next_payment_date",
            "fic_mis_date",
            "n_curr_payment_recd",
            "n_eop_bal",
            "v_amrt_repayment_type",
        )

        errors = {"rule1": [], "rule2": [], "rule3": [], "rule4": [], "rule5": [], "rule6": []}

        for record in cashflow_data:
            v_account_number = record["v_account_number"]
            d_maturity_date = record["d_maturity_date"]
            d_next_payment_date = record["d_next_payment_date"]
            fic_mis_date = record["fic_mis_date"]
            n_curr_payment_recd = record["n_curr_payment_recd"]
            n_eop_bal = record["n_eop_bal"]
            v_amrt_repayment_type = record["v_amrt_repayment_type"]

            # Rule 1
            if d_maturity_date < d_next_payment_date:
                errors["rule1"].append({
                    "v_account_number": v_account_number,
                    "error": f"Maturity date {d_maturity_date} must be greater than or equal to the next payment date {d_next_payment_date}.",
                })

            # Rule 2
            if d_next_payment_date > d_maturity_date:
                errors["rule2"].append({
                    "v_account_number": v_account_number,
                    "error": f"Next payment date {d_next_payment_date} must be less than or equal to maturity date {d_maturity_date}.",
                })

            # Rule 3
            if fic_mis_date >= d_maturity_date or fic_mis_date >= d_next_payment_date:
                errors["rule3"].append({
                    "v_account_number": v_account_number,
                    "error": f"Reporting Date {fic_mis_date} must be less than maturity date {d_maturity_date} and next payment date {d_next_payment_date}.",
                })

            # Rule 4
            if d_maturity_date == d_next_payment_date and n_curr_payment_recd < n_eop_bal:
                errors["rule4"].append({
                    "v_account_number": v_account_number,
                    "error": f"IF {d_maturity_date}={d_next_payment_date} then Current payment {n_curr_payment_recd} must be greater than or equal to end-of-period balance {n_eop_bal}.",
                })

            # Rule 5
            if d_maturity_date > d_next_payment_date and n_curr_payment_recd >= n_eop_bal:
                errors["rule5"].append({
                    "v_account_number": v_account_number,
                    "error": f"IF {d_maturity_date}>{d_next_payment_date} then Current payment {n_curr_payment_recd} must be less than end-of-period balance {n_eop_bal}.",
                })

            # Rule 6
            if v_amrt_repayment_type not in ["BULLET", "AMORTIZED"] or v_amrt_repayment_type is None:
                errors["rule6"].append({
                    "v_account_number": v_account_number,
                    "error": f"Repayment type must be 'BULLET' or 'AMORTIZED' but found '{v_amrt_repayment_type}'.",
                })

        if download:
            # Create Excel workbook
            workbook = openpyxl.Workbook()
            del workbook["Sheet"]

            for rule, rule_errors in errors.items():
                sheet = workbook.create_sheet(title=rule.upper())
                headers = ["#", "Account Number", "Error"]
                sheet.append(headers)

                for index, error in enumerate(rule_errors, start=1):
                    sheet.append([
                        index,
                        error["v_account_number"],
                        error["error"]
                    ])

                # Adjust column width
                for col_num, column in enumerate(headers, start=1):
                    sheet.column_dimensions[get_column_letter(col_num)].width = 25

            # Save the Excel file
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f"attachment; filename=cashflow_data_errors_{fic_mis_date}.xlsx"
            workbook.save(response)
            return response

        # Paginate errors
        paginated_errors = {}
        for rule, rule_errors in errors.items():
            paginator = Paginator(rule_errors, rows_per_page)
            page_number = int(request.GET.get(f"{rule}_page", 1))
            page_obj = paginator.get_page(page_number)
            paginated_errors[rule] = {
                "errors": list(page_obj),
                "total_pages": paginator.num_pages,
                "current_page": page_number,
                "total_records": paginator.count,
            }

        return JsonResponse({"paginated_errors": paginated_errors})

    return render(request, "operations/check_cashflow_data.html")





INSERT INTO `dim_function` (`function_name`, `description`) VALUES 
('perform_interpolation', 'Performs interpolation on the given data.'),
('run_vasicek_pit_PD_values', 'Calculate PIT PDs using Vasicek Model'),
('provision_matrix', 'Performs provision matrix on the historical data.'),
('run_lgd_calculation_dpd', 'Calculate  LGD using recovery rate'),
('run_frye_jacobs_pit_LGD_values', 'Calculate PIT LGD '),
('project_cash_flows', 'Projects future cash flows based on data.'),
('insert_fct_stage', 'Inserts records into the FCT Stage Determination.'),
('update_stage', 'Determines and updates the current stage of accounts.'),
('process_cooling_period_for_accounts', 'Processes the cooling period for accounts.'),
('update_stage_determination', 'Updates the stage determination logic.'),
('update_stage_determination_EAD_w_ACCR', 'Updates stage determination with accrued interest and EAD.'),
('update_stage_determination_accrued_interest_and_ead', 'Updates stage determination with accrued interest and EAD.'),
('update_eir_using_intrest_rate', 'Updates stage determination based on the effective interest rate.'),
('update_eir_using_cashflows', 'Updates stage determination based on the effective interest rate.'),
('update_lgd_for_stage_determination_term_structure', 'Updates loss given default for stage determination.'),
('update_lgd_for_stage_determination_term_structure_w_bands', 'Updates loss given default for stage determination.'),
('update_lgd_for_stage_determination_collateral', 'Updates loss given default for stage determination.'),
('calculate_pd_for_accounts', 'Calculates probability of default for accounts.'),
('insert_cash_flow_data', 'Inserts cash flow data into the system.'),
('update_financial_cash_flow', 'Updates financial cash flow records.'),
('calculate_discount_factors', 'Calculates discount factors for cash flows.'),
('calculate_ead_by_buckets', 'Calculates exposure at default for cash flows.'),
('update_cash_flow_with_pd_buckets', 'Updates cash flows with probability of default buckets.'),
('update_marginal_pd', 'Updates marginal probability of default calculations.'),
('cal_lgd_and_loss_rate_for_cash_flows_using_collateral', 'Calculates LGD using Collateral for cash flows.'),
('calculate_expected_cash_flow', 'Calculates expected cash flow projections.'),
('update_stage_determination_ead_with_cashflow_pv', 'Calculates exposure at default using cash flows pv.'),
('calculate_forward_loss_fields', 'Calculates forward-looking loss fields.'),
('calculate_cashflow_fields', 'Calculates various cash flow fields.'),
('populate_fct_reporting_lines', 'Populates reporting lines for FCT.'),
('calculate_ecl_based_on_method', 'Calculates expected credit loss based on selected method.'),
('update_reporting_lines_with_exchange_rate', 'Updates reporting lines with the  exchange rates.');

